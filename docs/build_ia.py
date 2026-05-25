"""PathWave Information Architecture (IA) Excel 빌더.

총 71 화면 (USER 26 / Provider 24 / Admin 21) 을 docs/ 폴더에 xlsx 로 출력.
- 시트 1: Cover (메타데이터 + 통계)
- 시트 2: IA (전체 화면 목록)
- 컬럼: NO | 구분 | 1depth | 2depth | 3depth | 4depth | 화면명 | 화면ID |
        GIT PR NO | Page | Tab | Layer Popup | Window Popup | Other |
        Native | Web | 진척율 | 기능 | 화면구성 | 비고
- 진척율, 합계는 Excel 수식 (COUNTIF / AVERAGE) — 하드코딩 금지
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = '/Users/m5pro16/Desktop/pathwave/docs/pathwave_information_architecture_2026-05-25.xlsx'

# ─── 컬럼 정의 ─────────────────────────────────────────────────────────────
COLS = [
    ('NO',            5),
    ('구분',          10),
    ('1depth',        14),
    ('2depth',        18),
    ('3depth',        18),
    ('4depth',        14),
    ('화면명',        22),
    ('화면ID',         9),
    ('GIT PR NO',     11),
    ('Page',           7),
    ('Tab',            6),
    ('Layer Popup',   12),
    ('Window Popup',  13),
    ('Other',          7),
    ('Native',         7),
    ('Web',            6),
    ('진척율',         8),
    ('기능',          55),
    ('화면구성',      30),
    ('비고',          24),
]

# ─── 데이터: (구분, 1depth, 2depth, 3depth, 4depth, 화면명, 화면ID, PR,
#           page,tab,layer,window,other, native,web, pct, 기능, 화면구성, 비고)
# Page/Tab/LayerPopup/WindowPopup/Other/Native/Web: True/False
# pct: int (0~100)
ROWS = [
    # ─── USER (mobile) ──────────────────────────────────────────────────
    ('USER', '인증', '스플래시', '', '', '스플래시', 'U-001', '',
     True, False, False, False, False, True, False, 100,
     '부팅 화면, 강제 업데이트 체크 (P10)', 'splash_screen.dart', ''),
    ('USER', '인증', '로그인', '', '', '로그인', 'U-002', '',
     True, False, False, False, False, True, False, 100,
     '이메일/소셜(카카오/네이버) 로그인', 'login_screen.dart', ''),
    ('USER', '인증', '회원가입', '', '', '회원가입', 'U-003', '#161',
     True, False, False, False, False, True, False, 100,
     '이메일 인증→비번→생년→약관 동의 단계', 'register_screen.dart', ''),
    ('USER', '인증', '회원가입', '약관 동의', '', '약관 동의', 'U-004', '#164',
     True, False, False, False, False, True, False, 100,
     'age14/terms_user/privacy_user/location 필수', 'consent_screen.dart',
     'C-2-4d (user/facility 분리)'),
    ('USER', '인증', '비밀번호 찾기', '', '', '비밀번호 찾기', 'U-005', '',
     True, False, False, False, False, True, False, 100,
     '이메일 코드 → 새 비번 설정', 'forgot_password_screen.dart', ''),
    ('USER', '홈', '홈', '', '', '홈', 'U-006', '',
     True, False, False, False, False, True, False, 100,
     '근처 매장 / 즐겨찾기 / 추천', 'home_screen.dart', ''),
    ('USER', '홈', 'WiFi 자동연결', '', '', 'WiFi 연결', 'U-007', '',
     True, False, False, False, False, True, False, 90,
     '비콘 감지 → .mobileconfig 설치 (1회 연결)',
     'wifi_connect_screen.dart', 'P16-b BLE 무중단 미구현 (실비콘 필요)'),
    ('USER', '검색', '매장 검색', '', '', '매장 검색', 'U-008', '',
     True, False, False, False, False, True, False, 100,
     '키워드 / 카테고리 필터', 'search_screen.dart', ''),
    ('USER', '매장', '매장 상세', '', '', '매장 상세', 'U-009', '',
     True, False, False, False, False, True, False, 100,
     '매장 정보 / 메뉴 / 스탬프 / 쿠폰 / 채팅 진입',
     'facility_screen.dart', ''),
    ('USER', '채팅', '채팅 목록', '', '', '채팅 목록', 'U-010', '',
     True, False, False, False, False, True, False, 100,
     '1:1 매장 채팅 목록', 'chat_list_screen.dart', ''),
    ('USER', '채팅', '채팅 상세', '', '', '채팅 상세', 'U-011', '',
     True, False, False, False, False, True, False, 100,
     '자동 번역 (DeepL) + 우상단 ⋮ 메뉴 (신고/차단)',
     'chat_detail_screen.dart', ''),
    ('USER', '채팅', '채팅 상세', '신고 시트', '', '신고 시트', 'U-012', '',
     False, False, True, False, False, True, False, 100,
     '사유 선택 + 상세 입력 → /api/chat/reports', '신고 모달', ''),
    ('USER', '채팅', '채팅 상세', '차단 확인', '', '차단 확인', 'U-013', '',
     False, False, True, False, False, True, False, 100,
     '매장 차단 확인 다이얼로그 → /api/chat/blocks', '차단 모달', ''),
    ('USER', '알림', '알림 목록', '', '', '알림', 'U-014', '',
     True, False, False, False, False, True, False, 100,
     '시스템 공지 / 매장 공지 / 약관 변경',
     'notifications_screen.dart', ''),
    ('USER', '마이페이지', '마이페이지', '', '', '마이페이지', 'U-015', '',
     True, False, False, False, False, True, False, 30,
     '/mypage 라우트 진입용 placeholder (홈의 마이 탭 사용 유도)',
     'mypage_screen.dart (37줄)',
     '실제 마이페이지 = 홈의 마이 탭. 직접 라우트는 redirect 안내'),
    ('USER', '마이페이지', '쿠폰함', '', '', '쿠폰함', 'U-016', '',
     True, False, False, False, False, True, False, 100,
     'status=active/used/expired 필터', 'coupons_screen.dart', ''),
    ('USER', '마이페이지', '스탬프', '', '', '스탬프', 'U-017', '',
     True, False, False, False, False, True, False, 100,
     '매장별 스탬프 적립 카드', 'stamps_screen.dart', ''),
    ('USER', '마이페이지', '즐겨찾기', '', '', '즐겨찾기', 'U-018', '',
     True, False, False, False, False, True, False, 100,
     '즐겨찾기 매장 목록', 'favorites_screen.dart', ''),
    ('USER', '마이페이지', '계정 삭제', '', '', '계정 삭제', 'U-019', '#181',
     True, False, False, False, False, True, False, 100,
     '비번 재확인 + 사유 → DELETE /api/auth/me (soft delete, anonymize)',
     'delete_account_screen.dart', 'Apple 5.1.1(v) 대응'),
    ('USER', '마이페이지', '보호자 초대', '', '', '보호자 초대', 'U-020', '',
     True, False, False, False, False, True, False, 100,
     '만 14 미만 부모 동의 흐름 (PR #47)',
     'parent_invite_screen.dart', ''),
    ('USER', '설정', '설정', '', '', '설정', 'U-021', '',
     True, False, False, False, False, True, False, 100,
     '알림 prefs / 언어 / 접근성', 'settings_screen.dart',
     'C-2-2 알림 카테고리 on/off'),
    ('USER', '설정', '비밀번호 변경', '', '', '비밀번호 변경', 'U-022', '',
     True, False, False, False, False, True, False, 100,
     '현재 비번 + 새 비번', 'change_password_screen.dart', ''),
    ('USER', '설정', '차단 목록', '', '', '차단 목록', 'U-023', '',
     True, False, False, False, False, True, False, 100,
     '차단한 매장 목록 + 해제', 'blocked_facilities_screen.dart', ''),
    ('USER', '설정', '약관 보기', '', '', '약관 보기', 'U-024', '#163',
     True, False, False, False, False, True, False, 100,
     'kind + version 드롭다운, 본문 보기',
     'policy_view_screen.dart', 'ko/en lang fallback'),
    ('USER', '고객지원', '문의 목록', '', '', '고객지원', 'U-025', '#169',
     True, False, False, False, False, True, False, 100,
     'FAQ 20행 + 1:1 문의 목록', 'support_screen.dart',
     'C-2-3 FAQ 시드 (user/provider × ko/en)'),
    ('USER', '고객지원', '문의 상세', '', '', '문의 상세', 'U-026', '',
     True, False, False, False, False, True, False, 100,
     '답변 채팅 + 자동 번역', 'support_detail_screen.dart', ''),

    # ─── Provider (provider-web) ────────────────────────────────────────
    ('Provider', '인증', '로그인', '', '', '로그인', 'P-001', '',
     True, False, False, False, False, False, True, 100,
     '이메일 로그인 + 게스트 둘러보기', 'Login.jsx',
     '게스트 모드 PR #66~#87'),
    ('Provider', '인증', '회원가입', '', '', '회원가입', 'P-002', '#161',
     True, False, False, False, False, False, True, 100,
     '사업자번호 / 담당자 / 약관 / 매장 정보 단계',
     'Signup.jsx', 'terms_facility/privacy_facility 필수'),
    ('Provider', '운영', '대시보드', '', '', '대시보드', 'P-003', '',
     True, False, False, False, False, False, True, 100,
     '매장 KPI 카드 (방문/스탬프/쿠폰/매출)', 'Dashboard.jsx',
     '녹색 #22C55E 포인트'),
    ('Provider', '매장', '매장 정보', '', '', '매장 정보', 'P-004', '',
     True, False, False, False, False, False, True, 100,
     '이름/주소/사진/시간/카테고리', 'StoreInfo.jsx',
     '1계정 1매장 정책'),
    ('Provider', '매장', '매장 정보', '사업자 정보', '', '사업자 정보 모달',
     'P-005', '',
     False, False, True, False, False, False, True, 100,
     '사업자등록증 / 통신판매업 정보',
     'BusinessInfoModal.jsx', ''),
    ('Provider', '운영', '스탬프 관리', '', '', '스탬프 관리', 'P-006', '',
     True, False, False, False, False, False, True, 100,
     '스탬프 정책 + 적립 이력', 'Stamps.jsx', ''),
    ('Provider', '운영', '스탬프 관리', '스탬프 발급', '', '스탬프 발급',
     'P-007', '',
     True, False, False, False, False, False, True, 100,
     '사용자 선택 + 결제 금액 입력', 'StampForm.jsx', ''),
    ('Provider', '운영', '쿠폰 관리', '', '', '쿠폰 관리', 'P-008', '',
     True, False, False, False, False, False, True, 100,
     '캠페인 목록 (발급/사용/잔여)', 'Coupons.jsx', ''),
    ('Provider', '운영', '쿠폰 관리', '쿠폰 발급', '', '쿠폰 발급',
     'P-009', '',
     True, False, False, False, False, False, True, 100,
     '대상 / 혜택 / 유효기간', 'CouponForm.jsx', ''),
    ('Provider', '채팅', '사용자 채팅', '', '', '사용자 채팅', 'P-010', '',
     True, False, False, False, False, False, True, 100,
     '1:1 자동 번역 (DeepL stub)', 'CustomerChat.jsx',
     'X1 cross 검증 통과 (PR #177)'),
    ('Provider', '알림', '알림 목록', '', '', '알림', 'P-011', '',
     True, False, False, False, False, False, True, 100,
     '매장 공지 발송 + 이력', 'Notifications.jsx', ''),
    ('Provider', '직원', '직원 관리', '', '', '직원 관리', 'P-012', '#175',
     True, False, False, False, False, False, True, 100,
     '초대 / 해지 / 권한', 'StaffManagement.jsx',
     'owner-only vs staff 권한 분리 검증'),
    ('Provider', '직원', '직원 관리', '초대 모달', '', '직원 초대 모달',
     'P-013', '#175',
     False, False, True, False, False, False, True, 100,
     '이메일 + role(staff/admin) → invite token 발급',
     'StaffManagement.jsx 내 모달',
     'invite_token 은 응답 비노출 (보안)'),
    ('Provider', '직원', '직원 관리', '작업 모달', '', '직원 작업 모달',
     'P-014', '',
     False, False, True, False, False, False, True, 100,
     '재발송 / 해지', '', ''),
    ('Provider', 'WiFi', 'WiFi 설정', '', '', 'WiFi 설정', 'P-015', '',
     True, False, False, False, False, False, True, 100,
     'SSID + 비번 (AES-256-GCM) + 비콘 매핑',
     'WifiSettings.jsx', 'P15a/P15b 완료'),
    ('Provider', '결제', '결제 관리', '', '', '결제 관리', 'P-016', '',
     True, False, False, False, False, False, True, 100,
     '카드 등록 + 영수증 다운로드', 'PaymentManagement.jsx',
     '토스 sandbox 자리만 (실키 단계2)'),
    ('Provider', '결제', '결제 관리', '카드 교체', '', '카드 교체 모달',
     'P-017', '',
     False, False, True, False, False, False, True, 100,
     '카드 교체 흐름', 'PaymentManagement 내 모달', ''),
    ('Provider', '결제', '구독', '', '', '구독', 'P-018', '',
     True, False, False, False, False, False, True, 100,
     '플랜 변경 / 해지 / 갱신', 'Subscriptions.jsx', ''),
    ('Provider', '회원', '회원 프로필', '', '', '회원 프로필', 'P-019', '',
     True, False, False, False, False, False, True, 100,
     '매장 단골 정보 + 활동 이력', 'MemberProfile.jsx', ''),
    ('Provider', '서비스', '서비스 신청', '', '', '서비스 신청',
     'P-020', '',
     True, False, False, False, False, False, True, 100,
     '비콘 설치 / 이전 / AS 신청', 'ServiceRequest.jsx', ''),
    ('Provider', '매장', '매장 목록', '', '', '매장 관리', 'P-021', '',
     True, False, False, False, False, False, True, 80,
     '내 매장 + (운영자 권한 시 다중 매장)', 'Facilities.jsx',
     '1계정 1매장 정책으로 표시 위주'),
    ('Provider', '설정', '설정', '', '', '설정', 'P-022', '',
     True, False, False, False, False, False, True, 100,
     '알림 prefs / 매장 옵션', 'Settings.jsx', ''),
    ('Provider', '고객지원', '문의', '', '', '고객지원', 'P-023', '',
     True, False, False, False, False, False, True, 100,
     '운영자 문의 + FAQ', 'Support.jsx', ''),
    ('Provider', '인증', '약관 보기', '', '', '약관 보기', 'P-024', '',
     True, False, False, False, False, False, True, 100,
     'kind+version 본문', 'PolicyView.jsx', ''),

    # ─── Admin (admin-web) ──────────────────────────────────────────────
    ('Admin', '인증', '로그인', '', '', '로그인', 'A-001', '',
     True, False, False, False, False, False, True, 100,
     '슈퍼어드민 로그인 (BOOTSTRAP_ADMIN env)', 'Login.jsx', ''),
    ('Admin', '메인', '대시보드', '', '', '대시보드', 'A-002', '',
     True, False, False, False, False, False, True, 100,
     '전사 KPI 카드 (users/facility/beacons/payments)',
     'Dashboard.jsx', '블루 #2563EB 포인트'),
    ('Admin', '메인', '비콘 관리', '', '', '비콘 관리', 'A-003', '',
     True, False, False, False, False, False, True, 100,
     'CSV 입고 / 목록 / 할당', 'Beacons.jsx',
     'P6-2 자동 검증 통과 (#173)'),
    ('Admin', '메인', '비콘 관리', '입고 모달', '', '비콘 입고 모달',
     'A-004', '',
     False, False, True, False, False, False, True, 100,
     'serial_no + uuid 입력 (CSV 또는 수동)', 'Beacons.jsx 내 모달', ''),
    ('Admin', '메인', '매장 승인', '', '', '매장 가입 승인', 'A-005', '',
     True, False, False, False, False, False, True, 100,
     'pending/verified/suspended 4종 액션',
     'Approvals.jsx', 'P6-3/P6-4 검증'),
    ('Admin', '운영', '배터리', '', '', '배터리 모니터링', 'A-006', '',
     True, False, False, False, False, False, True, 100,
     '저전력 비콘 알림 (low_threshold)', 'Battery.jsx', ''),
    ('Admin', '운영', '공지 작성', '', '', '시스템 공지', 'A-007', '',
     True, False, False, False, False, False, True, 100,
     'audience / lang_hint / push 옵션', 'Announcements.jsx', ''),
    ('Admin', '운영', '알림 검토', '', '', '알림 검토', 'A-008', '',
     True, False, False, False, False, False, True, 100,
     '발송 이력 + 대상 검토', 'Notifications.jsx', ''),
    ('Admin', '운영', '직원 모니터링', '', '', '직원 모니터링',
     'A-009', '#186',
     True, False, False, False, False, False, True, 100,
     '전체 staff_accounts + 매장/owner 매핑 + invitation_status + role 별 집계',
     'StaffMonitor.jsx',
     'GET /api/admin/staff/reports 신규 (D번들2)'),
    ('Admin', '운영', '채팅 모니터링', '', '', '채팅 모니터링',
     'A-010', '#186',
     True, False, False, False, False, False, True, 100,
     '전체 chat_rooms + 매장/사용자 + 메시지 수 + 마지막 활동 시각 (최대 500)',
     'ChatMonitor.jsx',
     'GET /api/admin/chat/rooms 신규 (D번들2)'),
    ('Admin', '운영', '신고 처리', '', '', '신고 처리', 'A-011', '',
     True, False, False, False, False, False, True, 100,
     '신고 큐 / 채팅 맥락 / 처리 액션', 'AbuseReports.jsx', ''),
    ('Admin', '결제·정책', '결제 관리', '', '', '결제 관리', 'A-012', '',
     True, False, False, False, False, False, True, 100,
     '결제 목록 + 환불 처리', 'Payments.jsx',
     'P6-9 환불 R2 deferred (실키 필요)'),
    ('Admin', '결제·정책', '약관 관리', '', '', '약관 관리',
     'A-013', '#162,#163',
     True, False, False, False, False, False, True, 100,
     '약관 목록 + multilang 발행', 'Policies.jsx', 'C-2-4b/c'),
    ('Admin', '결제·정책', '약관 관리', '약관 에디터', '', '약관 에디터',
     'A-014', '#163',
     False, False, True, False, False, False, True, 100,
     'ko/en 탭 split (신규) + 단일 lang 편집 (수정)',
     'PolicyEditor.jsx', 'C-2-4c'),
    ('Admin', '결제·정책', '쿠폰 통계', '', '', '쿠폰 통계',
     'A-015', '#186',
     True, False, False, False, False, False, True, 100,
     '쿠폰 발급/사용/활성/만료 카드 (백엔드 summary 사용)',
     'CouponStats.jsx',
     'GET /api/admin/coupons (?status=all|active|used|expired) 신규 (D번들2)'),
    ('Admin', '고객지원', '문의 목록', '', '', '고객지원', 'A-016', '',
     True, False, False, False, False, False, True, 100,
     '사용자 문의 검토', 'Support.jsx', ''),
    ('Admin', '고객지원', 'FAQ 관리', '', '', 'FAQ 관리', 'A-017', '#169',
     True, False, False, False, False, False, True, 100,
     'kind/lang/category CRUD', 'Faq.jsx',
     '시드 20행 자동 (C-2-3)'),
    ('Admin', '고객지원', '지원 통계', '', '', '고객지원 통계',
     'A-018', '',
     True, False, False, False, False, False, True, 100,
     '응답시간 / 카테고리 분포', 'SupportStats.jsx', ''),
    ('Admin', '시스템', '법인 정보', '', '', '법인 정보', 'A-019', '',
     True, False, False, False, False, False, True, 100,
     '사업자등록증 / 통신판매업 / 본점', 'CompanyInfo.jsx', ''),
    ('Admin', '시스템', '앱 버전', '', '', '앱 버전', 'A-020', '#180',
     True, False, False, False, False, False, True, 100,
     'iOS/Android min_supported/latest/store_url/force_message',
     'AppVersions.jsx', 'B-2d (C-2-1 운영 UI)'),
    ('Admin', '시스템', '다국어', '', '', '다국어 관리', 'A-021', '',
     True, False, False, False, False, False, True, 100,
     'i18n 키 관리 + DeepL 자동', 'Translations.jsx', ''),
    ('Admin', '시스템', '업종 카테고리', '', '', '업종 카테고리 관리',
     'A-027', '#193',
     True, False, False, False, False, False, True, 100,
     '국세청 100대 생활업종 시드 + 검색/그룹/상태 필터 + 신규/수정/비활성/완전삭제. 사장 가입 시 자유 입력 차단 (DB 파편화 방지).',
     'Categories.jsx',
     'GET /api/categories(공개) + admin CRUD. provider CategoryService 도 백엔드 fetch 로 전환 (polish)'),

    # ─── D 번들4-pre — 외부 AI 비용 모니터링 (2026-05-25) ─────────
    ('Admin', '시스템', 'AI 비용 모니터', '', '', 'AI 비용 모니터',
     'A-025', '#189',
     True, False, False, False, False, False, True, 100,
     '월 누적 USD/KRW + 임계점 진행률 바 + provider/operation 분류 + 활성 알림',
     'CostMonitor.jsx',
     'GET /api/admin/cost-monitor. 임계점 $100 ($1=₩1510.20)'),
    ('Admin', '시스템', 'AI 비용 모니터', '글로벌 알림', '', '글로벌 임계점 알림 모달',
     'A-026', '#189',
     False, False, True, False, False, False, True, 100,
     '슈퍼어드민 전용 critical 모달. 80%/100% 도달 시 자동 표시. 닫기=snooze (80%=24h / 100%=2h). 1분 polling',
     'CriticalAdminAlert.jsx',
     'GET /api/admin/critical-alerts + POST /alerts/{id}/dismiss. 가이드 docs/translation_cost_runaway_plan.md'),

    # ─── 누락 — 구현 필요 항목 (2026-05-25 IA 정확도 감사 후 추가) ────
    ('USER', '매장', '매장 상세', '메뉴 자동 번역', '', '메뉴 자동 번역 보기',
     'U-027', '#192',
     False, False, False, False, True, True, False, 100,
     '매장 상세 안 메뉴 섹션 — Localizations.localeOf 자동 lang + 백엔드 fallback (cache/translated/fallback_blocked) + 가격 KRW 보존 + 자동번역 배지',
     'facility_screen.dart _buildMenu',
     'C-4 USP. lang 변경 시 didChangeDependencies 로 재fetch. 가격은 환산/단위변경 없이 원본 그대로'),

    ('Provider', '매장', '매장 다국어', '', '', '매장 다국어 관리',
     'P-025', '#188',
     True, False, False, False, False, False, True, 100,
     '6개 언어 탭 (en/ja/zh/zh-TW/fr/th) + 수동 편집 + 자동 번역 (force 옵션) + 삭제',
     'StoreTranslations.jsx',
     '백엔드 /api/facilities/{fid}/translations[/auto] 활용 (D번들3-B)'),

    ('Provider', '매장', '매장 메뉴 OCR', '', '', '매장 메뉴 사진 OCR + 자동번역',
     'P-026', '#190,#191',
     True, False, False, False, False, False, True, 100,
     '백엔드 + provider UI 완료. 파일 picker → OCR → 인라인 표 (수정/추가/삭제 + 정렬) + replace 옵션. 가격 KRW 안내문.',
     'MenuManagement.jsx + MenuService.js',
     'C-4 USP. GCV provider (월 1000장 무료, 키 없으면 stub). mobile USER 화면 U-027 별도 (4-c)'),

    ('Admin', '운영', '회원 관리', '', '', '회원 관리 (사용자 조회)',
     'A-022', '#187',
     True, False, False, False, False, False, True, 100,
     '검색 + 필터(status/provider) + 페이지네이션 + 신고/채팅 카운트 + 강제 탈퇴 모달',
     'Users.jsx',
     'GET /api/admin/users + POST /users/{id}/force-delete 신규 (D번들3-A). Apple 5.1.1(v) 준수 soft-delete + 이메일 anonymize'),

    ('Admin', '시스템', '시스템 환경 점검', '', '', '시스템 환경 점검',
     'A-023', '#185',
     True, False, False, False, False, False, True, 100,
     '외부 키 (Firebase/DeepL/Anthropic/SendGrid/Toss/Sentry/Maps) 설정 상태 + 모드 (live/stub/missing) 카드 그리드',
     'SystemHealth.jsx',
     'GET /api/admin/system/health + admin 페이지 신규 (D번들1)'),

    ('Admin', '시스템', '운영자 비밀번호 변경', '', '', '운영자 비밀번호 변경',
     'A-024', '#185',
     False, False, True, False, False, False, True, 100,
     '슈퍼어드민 본인 비밀번호 변경 (사이드바 키 아이콘 → 모달, 현재 비번 재확인)',
     'ChangePasswordModal.jsx',
     'POST /api/admin/change-password + 모달 신규 (D번들1). bcrypt + 비번 정책 검증 + rate limit'),
]


# ─── 색상 / 스타일 ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', start_color='2C2C38', end_color='2C2C38')
HEADER_FONT = Font(name='맑은 고딕',size=10, bold=True, color='FFFFFF')
USER_FILL     = PatternFill('solid', start_color='DCE7F5', end_color='DCE7F5')
PROVIDER_FILL = PatternFill('solid', start_color='DDF1DC', end_color='DDF1DC')
ADMIN_FILL    = PatternFill('solid', start_color='F8E6CB', end_color='F8E6CB')
PCT_FILL_FULL  = PatternFill('solid', start_color='B7E5B5', end_color='B7E5B5')
PCT_FILL_PART  = PatternFill('solid', start_color='FFEAB0', end_color='FFEAB0')
TOTAL_FILL = PatternFill('solid', start_color='3A3A48', end_color='3A3A48')
TOTAL_FONT = Font(name='맑은 고딕',size=10, bold=True, color='FFFFFF')

thin = Side(border_style='thin', color='AEB1BC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
DEFAULT_FONT = Font(name='맑은 고딕',size=10)


def fill_for(category: str):
    return {'USER': USER_FILL, 'Provider': PROVIDER_FILL,
            'Admin': ADMIN_FILL}.get(category)


# ─── Cover sheet ───────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.active
    ws.title = 'Cover'
    ws.sheet_view.showGridLines = False

    ws['B2'] = 'PathWave — Information Architecture (IA)'
    ws['B2'].font = Font(name='맑은 고딕',size=18, bold=True)
    ws['B3'] = '3 콘솔 (USER / Provider / Admin) 전체 화면 인포메이션 아키텍처'
    ws['B3'].font = Font(name='맑은 고딕',size=11, color='5C6170')

    meta = [
        ('작성일',       '2026-05-25'),
        ('총 화면 수',    f'{len(ROWS)} 화면'),
        ('USER 화면',     sum(1 for r in ROWS if r[0] == 'USER')),
        ('Provider 화면', sum(1 for r in ROWS if r[0] == 'Provider')),
        ('Admin 화면',    sum(1 for r in ROWS if r[0] == 'Admin')),
        ('평균 진척율',   "='IA'!Q" + str(len(ROWS) + 2)),
        ('전체 PR 진행',  '#161~#181 모두 머지 (2026-05-25 기준)'),
        ('단계',          '출시 5단계 중 1단계 완료 — 다음 = 2단계 (외부 서비스 신청)'),
    ]
    for i, (k, v) in enumerate(meta, start=5):
        ws[f'B{i}'] = k
        ws[f'B{i}'].font = Font(name='맑은 고딕',size=10, bold=True)
        ws[f'B{i}'].fill = PatternFill('solid', start_color='EFF2F7',
                                       end_color='EFF2F7')
        ws[f'C{i}'] = v
        ws[f'C{i}'].font = DEFAULT_FONT
        if k == '평균 진척율':
            ws[f'C{i}'].number_format = '0.0%'

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 55

    # 범례
    base = len(meta) + 7
    ws[f'B{base}'] = '범례 / 컬럼 의미'
    ws[f'B{base}'].font = Font(name='맑은 고딕',size=12, bold=True)
    notes = [
        ('구분',         'USER (mobile) / Provider (provider-web) / Admin (admin-web)'),
        ('화면ID',       'U-001 / P-001 / A-001 형식. 본 문서 내 고유 식별자.'),
        ('GIT PR NO',    '해당 화면을 도입/수정한 PR 번호 (있는 경우만).'),
        ('Page',         '페이지 구성 화면인지 — 독립된 라우트/URL 한 페이지.'),
        ('Tab',          '화면에 탭인지 — 같은 페이지 안 탭 단위로 전환되는지.'),
        ('Layer Popup',  '화면에 레이어 노출되는지 — 오버레이 모달 (동일 라우트, 배경 딤드).'),
        ('Window Popup', '윈도우 팝업인지 — OS 윈도우/새 창 단위로 띄우는지.'),
        ('Other',        '기타 알럿 등 — 토스트 / 시트 / 인라인 메시지 등.'),
        ('Native',       '네이티브 구현 — mobile (Flutter / iOS·Android).'),
        ('Web',          '웹페이지 구현 — provider-web / admin-web (React + Vite).'),
        ('진척율',       '1단계 (local + stub) 기준. R2 (실키), R3 (심의 전) 단계에서 재조정.'),
        ('비고',         'P16-b 미구현 / C-4 보류 / 보안 제약 등 특기 사항.'),
    ]
    for i, (k, v) in enumerate(notes, start=base + 2):
        ws[f'B{i}'] = k
        ws[f'B{i}'].font = Font(name='맑은 고딕',size=10, bold=True,
                                color='2563EB')
        ws[f'C{i}'] = v
        ws[f'C{i}'].font = DEFAULT_FONT
        ws[f'C{i}'].alignment = LEFT_WRAP

    # ─── 지속 관리 가이드 ───
    mb = base + len(notes) + 4
    ws[f'B{mb}'] = '지속 관리 가이드 (실제 진척율 유지)'
    ws[f'B{mb}'].font = Font(name='맑은 고딕', size=12, bold=True)
    guide = [
        ('1. 본 파일 빌드',     '/tmp/build_ia.py 또는 docs/build_ia.py 로 재생성. ROWS 리스트 수정 후 재실행.'),
        ('2. PR 머지 후 갱신',  'PR 머지 시 IA 영향이 있으면 해당 행의 진척율 / 비고 / GIT PR NO 컬럼 동시 업데이트.'),
        ('3. placeholder 식별', '코드에 "// 엔드포인트 미구현" / "TODO" / 짧은 라인수 (<150) 는 진척율 의심.'),
        ('4. 백엔드 ↔ UI 매핑', '백엔드 API 가 있는데 UI 가 없으면 진척율 ≤30%. 둘 다 있으면 ≥80%.'),
        ('5. R2/R3 단계 진입',  '단계 3 (실 키) / 4 (심의 직전) 진입 시 IA 재감사 → 진척율 재조정.'),
        ('6. 누락 발견 시',     '구현 필요 화면이 IA 에 없으면 즉시 신규 행 추가 (진척율 0% 로).'),
        ('7. 보류 (v1 제외)',  '외국인 결제 / 면세 신청 등은 v1 스코프 제외. IA 에서도 제외.'),
    ]
    for i, (k, v) in enumerate(guide, start=mb + 2):
        ws[f'B{i}'] = k
        ws[f'B{i}'].font = Font(name='맑은 고딕', size=10, bold=True,
                                color='22C55E')
        ws[f'C{i}'] = v
        ws[f'C{i}'].font = DEFAULT_FONT
        ws[f'C{i}'].alignment = LEFT_WRAP

    ws.row_dimensions[2].height = 28


# ─── IA sheet ──────────────────────────────────────────────────────────────
def build_ia(wb):
    ws = wb.create_sheet('IA')

    # Header
    for col_idx, (name, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    for i, row in enumerate(ROWS, start=2):
        (cat, d1, d2, d3, d4, name, sid, pr,
         page, tab, layer, window, other, native, web,
         pct, func, ui, note) = row

        # NO
        c = ws.cell(row=i, column=1, value=i - 1)
        c.alignment = CENTER

        # 구분
        c = ws.cell(row=i, column=2, value=cat)
        c.fill = fill_for(cat)
        c.alignment = CENTER
        c.font = Font(name='맑은 고딕',size=10, bold=True)

        # depths
        ws.cell(row=i, column=3, value=d1).alignment = CENTER
        ws.cell(row=i, column=4, value=d2).alignment = LEFT_WRAP
        ws.cell(row=i, column=5, value=d3).alignment = LEFT_WRAP
        ws.cell(row=i, column=6, value=d4).alignment = LEFT_WRAP

        # 화면명
        c = ws.cell(row=i, column=7, value=name)
        c.font = Font(name='맑은 고딕',size=10, bold=True)
        c.alignment = LEFT_WRAP

        # 화면ID, PR
        ws.cell(row=i, column=8, value=sid).alignment = CENTER
        ws.cell(row=i, column=9, value=pr).alignment = CENTER

        # 체크 컬럼 (Page/Tab/Layer/Window/Other/Native/Web)
        for idx, flag in enumerate([page, tab, layer, window, other,
                                    native, web], start=10):
            v = '●' if flag else ''
            c = ws.cell(row=i, column=idx, value=v)
            c.alignment = CENTER
            if flag and idx in (10, 11, 12, 13, 14):
                c.font = Font(name='맑은 고딕',size=12, bold=True,
                              color='2563EB')
            elif flag and idx in (15, 16):
                c.font = Font(name='맑은 고딕',size=12, bold=True,
                              color='22C55E')

        # 진척율 (cell value 0~1 로 저장, format %)
        c = ws.cell(row=i, column=17, value=pct / 100)
        c.number_format = '0%'
        c.alignment = CENTER
        c.fill = PCT_FILL_FULL if pct == 100 else PCT_FILL_PART

        # 기능 / 화면구성 / 비고
        for idx, val in enumerate([func, ui, note], start=18):
            c = ws.cell(row=i, column=idx, value=val)
            c.alignment = LEFT_WRAP
            c.font = DEFAULT_FONT

        # 행 전체 테두리
        for col_idx in range(1, len(COLS) + 1):
            ws.cell(row=i, column=col_idx).border = BORDER

        # 행 높이
        ws.row_dimensions[i].height = 32

    # 합계 행
    total_row = len(ROWS) + 2
    ws.cell(row=total_row, column=1, value='합계').alignment = CENTER
    for col_idx in range(1, len(COLS) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # COUNTIF 수식 (Page/Tab/Layer/Window/Other/Native/Web 의 ● 갯수)
    data_start = 2
    data_end   = len(ROWS) + 1
    for col_idx in range(10, 17):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx).value = (
            f'=COUNTIF({col_letter}{data_start}:{col_letter}{data_end},"●")'
        )

    # 평균 진척율 (Q열 / col 17)
    pct_col_letter = get_column_letter(17)
    ws.cell(row=total_row, column=17).value = (
        f'=AVERAGE({pct_col_letter}{data_start}:{pct_col_letter}{data_end})'
    )
    ws.cell(row=total_row, column=17).number_format = '0.0%'

    # 좌측 라벨 (구분 컬럼) — USER/Provider/Admin 갯수 별도 표시
    ws.cell(row=total_row, column=2).value = (
        f'=COUNTIF(B{data_start}:B{data_end},"USER")&"+"&'
        f'COUNTIF(B{data_start}:B{data_end},"Provider")&"+"&'
        f'COUNTIF(B{data_start}:B{data_end},"Admin")'
    )

    # 1depth ~ 4depth, 화면명, 화면ID, GIT PR, 기능, 화면구성, 비고 빈 칸
    ws.row_dimensions[total_row].height = 28

    # Freeze panes (헤더 + NO/구분 컬럼 유지)
    ws.freeze_panes = 'C2'


def main():
    wb = Workbook()
    build_cover(wb)
    build_ia(wb)
    wb.save(OUT)
    print(f'wrote {OUT}')
    print(f'total rows: {len(ROWS)}')


if __name__ == '__main__':
    main()
