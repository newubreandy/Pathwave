"""PathWave 페르소나 통합 테스트 시나리오 — xlsx 빌더.

사용
----
    /Users/m5pro16/Desktop/pathwave/venv/bin/python docs/build_test_scenarios.py

산출
----
    docs/pathwave_persona_test_scenarios_2026-05-26.xlsx

관리 정책
--------
- IA 와 동일 라이프사이클. 화면ID 가 IA 와 매칭되도록 유지.
- PR 머지 시 변경 영향이 있으면 본 스크립트 ROWS 갱신 + xlsx 재빌드.
- 사용자(테스터) 가 xlsx 에서 직접 결과 컬럼 (Galaxy S24 / iPhone 16 Pro / Safari / Chrome /
  Pass 여부) 을 채우는 형태.

시트 구성
--------
1. Cover            — 작성일 / 페르소나 정의 / 결과 기호 / 컬럼 의미
2. Scenarios        — 17 컬럼 매트릭스 (사용자 요청 양식)
3. Persona-Index    — 페르소나별 시나리오 count

페르소나 14명
- P1   외국인 관광객 (mobile, lang=ja → en fallback)
- P2   한국인 일반 사용자 (mobile)
- P3   소규모 매장 사장 (provider-web)
- P4   중대형 매장 사장 + 직원 보유 (provider-web)
- P5   매장 직원 (provider-web, sub_type=staff)
- P6   슈퍼어드민 (admin-web)
- P7   친구 초대 발급자 (mobile, user → user)
- P8   친구 초대로 가입한 신규 사용자 (mobile, invitation_code)
- P9   만 14세 미만 + 보호자 동의 (mobile, parent flow)
- P10  소셜 로그인 사용자 (Kakao / Naver / Apple / Google)
- P11  비밀번호 재설정 (forgot/reset)
- P12  푸시 토큰 등록 (FCM/APNS)
- P13  1:1 문의 (Support ticket)
- P14  알림 발송 승인 (provider → admin)
"""
from __future__ import annotations

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ════════════════════════════════════════════════════════════════════════════
#  컬럼 정의 (사용자 요청 양식 그대로)
# ════════════════════════════════════════════════════════════════════════════
COLUMNS = [
    ('No',                 6),
    ('화면ID',              10),
    ('화면명',              22),
    ('구분',                10),   # USER / Provider / Admin
    ('테스트 수행 메뉴/화면\n(입력/선택/저장/조회)', 28),
    ('가이드/진입',          26),
    ('테스트 수행 내용',      40),
    ('테스트 데이터\n(상품번호/회원 아이디)', 24),
    ('테스트 결과 화면/메뉴\n(조회/확인)',  24),
    ('테스트 기대결과',       38),
    ('Galaxy S24',           7),
    ('iPhone 16 Pro',        9),
    ('Safari',               7),
    ('Chrome',               7),
    ('테스트결과\n(Pass여부)', 10),
    ('비고\n(결합내용 등)',   24),
    ('테스터\n(기대결과 확인)', 14),
]


# ════════════════════════════════════════════════════════════════════════════
#  페르소나 정의 (Cover 시트용)
# ════════════════════════════════════════════════════════════════════════════
PERSONAS = [
    ('P1',  '외국인 관광객',          'mobile (iOS) / 일본어 (lang=ja)',
     'i18n en fallback, 메뉴 자동번역, 채팅 자동번역'),
    ('P2',  '한국인 일반 사용자',      'mobile (Android) / 한국어',
     '표준 가입, BLE 스캔, 스탬프, 쿠폰, 1:1 채팅'),
    ('P3',  '소규모 매장 사장 (1인)',  'provider-web (Chrome/Safari)',
     '가입 + 운영자 승인 + 1계정 1매장 + 비콘 claim + WiFi 등록'),
    ('P4',  '중대형 매장 사장',        'provider-web',
     'P3 + 직원 초대 + 권한 분리 + 정산 + 통계'),
    ('P5',  '매장 직원',              'provider-web (초대 가입)',
     '초대 코드 가입 + 사장 권한 일부 + QR 스캔'),
    ('P6',  '슈퍼어드민',             'admin-web (Chrome)',
     '비콘 CSV / 매장 승인 / 약관 / 신고 처리 / 비용 모니터'),
    ('P7',  '친구 초대 발급자',        'mobile (user → user)',
     '초대 코드 발급 + 외부 채널 공유 + 친구 가입 확인'),
    ('P8',  '친구 초대 가입 사용자',    'mobile (invitation_code)',
     '코드 입력 → 가입 → 초대자와 연결'),
    ('P9',  '만 14세 미만 + 보호자',   'mobile (parent flow)',
     'birth_year < 14 → 보호자 동의 → 가입 완료'),
    ('P10', '소셜 로그인 사용자',      'mobile (Kakao/Naver/Apple/Google)',
     'OAuth → JWT 발급 → 약관 동의 흐름'),
    ('P11', '비밀번호 재설정',        'mobile / provider-web',
     'forgot-password → 이메일 링크 → reset-password'),
    ('P12', '푸시 토큰 등록',         'mobile (FCM/APNS)',
     '권한 동의 → 토큰 등록 → 매장 공지 수신'),
    ('P13', '1:1 문의 (Support)',    'mobile',
     'support ticket 작성 → 슈퍼어드민 답변 → 양방향 메시지'),
    ('P14', '알림 발송 승인 흐름',     'provider → admin → 사용자',
     '사장 발송 요청 → 슈퍼어드민 승인/거절 → dispatch'),
]


# ════════════════════════════════════════════════════════════════════════════
#  시나리오 데이터
#  필드: (no, 화면ID, 화면명, 구분, 메뉴/화면, 가이드/진입, 테스트 수행 내용,
#         테스트 데이터, 결과 화면/메뉴, 기대결과,
#         g24, iph16, safari, chrome, pass, 비고, 테스터)
#  g24/iph16/safari/chrome:  '●' = 대상,  ''  = 비대상
#  pass:  ''(미실행) / 'Pass' / 'Fail'
# ════════════════════════════════════════════════════════════════════════════
ROWS = []
_no = [0]
def add(persona: str, sid: str, sname: str, kind: str,
        menu: str, guide: str, action: str, data: str,
        result_screen: str, expected: str,
        g24='●', iph16='●', safari='', chrome='', notes=''):
    """비고 컬럼은 항상 'P{N}: ...' 형식. COUNTIFS 매칭 정확성 위해."""
    _no[0] += 1
    note_value = f'{persona}: {notes}' if notes else f'{persona}: '
    ROWS.append((_no[0], sid, sname, kind, menu, guide, action, data,
                 result_screen, expected, g24, iph16, safari, chrome, '',
                 note_value, ''))


# ─── P1 외국인 관광객 (mobile, lang=ja) ───────────────────────────────
add('P1', 'U-001', '스플래시', 'USER',
    '앱 첫 실행 — 디바이스 lang=ja',
    '앱 아이콘 터치',
    'OS 언어 ja 감지 → /api/i18n/ja 응답 (en fallback) → 로그인/온보딩 화면',
    '디바이스 ja 설정',
    '로그인 화면 (영문 UI)',
    'UI 텍스트가 영문 fallback 으로 표시. 한국어 잔존 X',
    iph16='●', g24='●', safari='', chrome='')
add('P1', 'U-003', '회원가입', 'USER',
    '이메일 + 코드 + 약관',
    '로그인 화면 → "Sign up"',
    '이메일 입력 → 코드 발송 → 코드 검증 → 비번 + 약관 동의',
    'jp_tour@pathwave.test',
    '약관 모달 (en 본문)',
    '/api/policies?sub_type=user&lang=ja → en fallback. terms_user/privacy_user 필수')
add('P1', 'U-007', 'WiFi 자동 연결', 'USER',
    '매장 입장 후 BLE 감지',
    '매장 입장 (비콘 신호 도달)',
    'BLE 스캔 → handshake → .mobileconfig 자동 설치 (iOS) → WiFi 연결',
    '매장 비콘 1개 (mock 가능)',
    'WiFi 연결됨 알림',
    'POST /api/beacon/handshake → 200 + wifi profile. 무중단 핸드오프는 P16-b (out of scope v1)')
add('P1', 'U-008', '매장 상세 — 메뉴 자동 번역', 'USER',
    '매장 상세 안 "Menu" 섹션',
    '매장 진입 (BLE 또는 검색)',
    '한국어 메뉴 OCR 결과 → 일본어 자동 번역 표시 (가격은 KRW 원본)',
    'facility_id 매장 메뉴 등록됨',
    '메뉴 항목 ja 표시 + 가격 ₩',
    'GET /api/facilities/{fid}/menu?lang=ja → source=translated. price 환산 X (KRW 그대로)')
add('P1', 'U-009', '채팅 자동 번역', 'USER',
    '매장과 1:1 채팅',
    '매장 상세 → "Chat"',
    '일본어 입력 → 사장 화면에 ko 번역+원문. 사장 ko 답변 → 사용자에 ja 번역+원문',
    'jp_tour ↔ shop1',
    '양방향 채팅 자동 번역',
    'tests/test_chat_translation.py 참조. DeepL stub 또는 실키')
add('P1', 'U-013', '푸시 알림 수신 (en)', 'USER',
    '매장 공지 수신',
    '매장 공지 발송 후',
    'FCM 푸시 본문 en (lang fallback). 한국어 단말은 ko 동시 발송 확인',
    '매장 공지 1건',
    '푸시 알림 (en 본문)',
    'P8c 푸시 번역 회귀 (test_push_translation)')

# ─── P2 한국인 일반 사용자 (mobile, ko) ────────────────────────────
add('P2', 'U-003', '회원가입 ko', 'USER',
    '이메일 + 코드 + 약관 ko',
    '로그인 → "회원가입"',
    '이메일 → 코드 발송 → 코드 검증 → 비번 + 약관 한국어 본문 동의',
    'kr_user@pathwave.test',
    '약관 모달 (ko 본문)',
    'terms_user/privacy_user 한국어. 가입 200 + access_token')
add('P2', 'U-006', '홈 — 근처 매장', 'USER',
    '홈 화면 + 위치 권한 허용',
    '로그인 후 홈',
    '근처 매장 (위치 기반) / 즐겨찾기 / 최근 방문',
    '서울 강남구 위치',
    '매장 카드 N개',
    'GET /api/search/facilities?lat=..&lon=..')
add('P2', 'U-010', '스탬프 적립', 'USER',
    '매장 직원이 QR 스캔',
    '매장 방문 후 QR 발급',
    '쿠폰 QR 표시 → 직원 스캔 → 스탬프 적립',
    'kr_user QR + 매장 직원',
    '스탬프 카드 +1',
    '매장 스탬프 정책 적용. 1일 1회 등')
add('P2', 'U-011', '쿠폰 사용', 'USER',
    '내 쿠폰 → 사용',
    '마이 → 쿠폰',
    '쿠폰 선택 → QR 표시 → 직원 스캔 → 사용 처리',
    '발급된 쿠폰 1건',
    '쿠폰 used=true',
    'POST /api/coupons/{id}/use. 만료/중복 차단 확인')
add('P2', 'U-016', '계정 삭제', 'USER',
    '설정 → 계정 삭제',
    '마이 → 설정',
    '"계정 삭제" → 비번 재확인 → 사유 → 확인',
    'kr_user / 본인 비번',
    '즉시 로그아웃',
    'DELETE /api/auth/me. 이메일 anonymize + push 토큰 삭제')

# ─── P3 소규모 매장 사장 (provider-web) ────────────────────────────
add('P3', 'P-001', '로그인', 'Provider',
    '로그인 화면',
    'provider-web /login',
    '이메일 + 비번 입력 → 로그인',
    'shop1@pathwave.test / Shop1234!',
    '대시보드 진입',
    '미인증 매장은 pending 화면. verified 매장은 dashboard',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-002', '가입', 'Provider',
    '회원가입',
    '/signup',
    '이메일/사업자번호/약관(terms_facility/privacy_facility) → 201 pending',
    '신규 사업자번호',
    '"승인 대기" 안내',
    'POST /api/facility/register. status=pending → admin verify 대기',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-004', '매장 정보 수정', 'Provider',
    '/dashboard/store',
    '대시보드 → 매장 정보',
    '매장명/주소/사진/운영시간 변경 → 저장',
    'fid 1',
    '저장 토스트',
    'PATCH /api/facilities/{fid}',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-005', '비콘 claim', 'Provider',
    '/dashboard/store → 비콘',
    'admin 이 입고한 SN 입력',
    'SN 입력 + role(wifi/cashier) 선택 → claim',
    'SN: PW-BCN-001 (admin 입고)',
    '비콘 active 표시',
    'POST /api/facilities/{fid}/claim-beacon',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-009', 'WiFi 프로파일 등록', 'Provider',
    '/dashboard/wifi',
    '대시보드 → 와이파이',
    'SSID + 비번 입력 → 저장 (AES-256-GCM 암호화 저장)',
    'SSID/비번',
    'WiFi 목록 +1',
    'POST /api/beacon/wifi. DB password 평문 X 검증',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-014', '메뉴 OCR 업로드', 'Provider',
    '/dashboard/menu',
    '메뉴 관리 → 사진 업로드',
    '메뉴판 이미지 (jpg/png) 업로드 → OCR 결과 인라인 표 확인 + 수정',
    '메뉴판 이미지',
    '메뉴 항목 표 (KRW 가격)',
    'POST /api/facilities/{fid}/menu/upload. 가격 KRW 강제',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-015', '매장 다국어 관리', 'Provider',
    '/dashboard/store-translations',
    '대시보드 → 매장 다국어',
    '영어 탭 → 매장명/주소/설명 입력 → 저장. 또는 "자동 번역"',
    'fid 1, lang=en',
    '저장 + en 탭 ● 표시',
    'PUT /api/facilities/{fid}/translations/en',
    g24='', iph16='', safari='●', chrome='●')
add('P3', 'P-012', '구독 결제', 'Provider',
    '/dashboard/payments',
    '결제 관리',
    '플랜 선택 → 카드 등록 → 토스 결제',
    '토스 sandbox',
    'subscriptions row',
    'POST /api/billing/subscriptions. R2 (실키) 단계',
    g24='', iph16='', safari='●', chrome='●')

# ─── P4 중대형 매장 사장 — 직원 보유 ─────────────────────────────
add('P4', 'P-019', '직원 초대 코드 발급', 'Provider',
    '/dashboard/staff → "초대"',
    '직원 관리',
    '이메일 + role(staff/admin) 입력 → 초대 발급',
    'staff1@pathwave.test',
    '초대 목록 +1, 토큰 발급',
    'POST /api/staff/invite. invite_token API 응답 비노출 (DB+이메일)',
    g24='', iph16='', safari='●', chrome='●')
add('P4', 'P-019', '직원 해지', 'Provider',
    '/dashboard/staff → "해지"',
    '직원 카드 → 해지',
    'pending 초대 revoke 또는 active staff 해지',
    'invitation_id 1',
    '직원 상태 revoked',
    'DELETE /api/staff/{iid}',
    g24='', iph16='', safari='●', chrome='●')

# ─── P5 매장 직원 ───────────────────────────────────────────────
add('P5', 'U-003', '직원 가입 (초대 코드)', 'Provider',
    '/signup?code=...',
    '이메일 받은 초대 링크',
    '코드 검증 → 비번 + 약관 동의 → 직원 계정 생성',
    'invite_token',
    '직원 로그인 가능',
    'POST /api/staff/accept. terms_facility/privacy_facility 필수',
    g24='', iph16='', safari='●', chrome='●')
add('P5', 'P-001', '직원 로그인', 'Provider',
    '/login (staff)',
    '이메일 + 비번',
    '직원 토큰 발급 + 대시보드 진입 (권한 제한)',
    'staff1@pathwave.test',
    '대시보드 (권한 제한)',
    'POST /api/staff/login',
    g24='', iph16='', safari='●', chrome='●')
add('P5', 'P-014', '직원 권한 — QR 스캔 가능', 'Provider',
    '/dashboard/stamps → 스캔',
    '스탬프 적립 화면',
    'QR 스캔 → 스탬프/쿠폰 처리',
    'user QR',
    '적립/사용 처리',
    'staff role 허용 endpoint',
    g24='', iph16='', safari='●', chrome='●')
add('P5', 'P-009', '직원 권한 — WiFi 등록 차단', 'Provider',
    '/dashboard/wifi (staff)',
    '직원 토큰으로 WiFi 등록 시도',
    'WiFi POST 시도',
    'WiFi 신규 SSID',
    '403 권한 거부',
    'POST /api/beacon/wifi 는 owner/admin 만',
    g24='', iph16='', safari='●', chrome='●')

# ─── P6 슈퍼어드민 ───────────────────────────────────────────────
add('P6', 'A-001', '슈퍼어드민 로그인', 'Admin',
    '/login',
    'admin-web /login',
    '이메일 + 비번 → 대시보드',
    'admin@pathwave.kr',
    '대시보드 진입',
    'POST /api/admin/login',
    g24='', iph16='', safari='●', chrome='●')
add('P6', 'A-003', '비콘 CSV 입고', 'Admin',
    '/dashboard/beacons',
    '비콘 → "CSV 입고"',
    'serial_no + uuid 배열 업로드',
    '3건 (PW-BCN-001~003)',
    'count=3 / status=inventory',
    'POST /api/admin/beacons/import',
    g24='', iph16='', safari='●', chrome='●')
add('P6', 'A-005', '매장 가입 승인', 'Admin',
    '/dashboard/approvals',
    '가입 신청 목록',
    'pending 행 → "승인"',
    'pending facility 1건',
    'status=verified',
    'POST /api/admin/facility-accounts/{aid}/verify',
    g24='', iph16='', safari='●', chrome='●')
add('P6', 'A-013', '약관 새 버전 발행 (ko+en)', 'Admin',
    '/dashboard/policies',
    '약관 → "새 버전 작성"',
    'kind/version/effective_at + ko/en 본문 동시 입력 → 발행',
    'terms_user v0.2',
    'pending 2 row (ko+en)',
    'POST /api/admin/policies/multilang',
    g24='', iph16='', safari='●', chrome='●')
add('P6', 'A-022', '회원 관리 → 강제 탈퇴', 'Admin',
    '/dashboard/users',
    '회원 관리 → 행 → "강제 탈퇴"',
    '사유 입력 → 확인',
    '사용자 id 1',
    'deleted_at 설정, 이메일 anonymize',
    'POST /api/admin/users/{uid}/force-delete',
    g24='', iph16='', safari='●', chrome='●')
add('P6', 'A-025', 'AI 비용 모니터', 'Admin',
    '/dashboard/cost-monitor',
    '시스템 → AI 비용 모니터',
    '월 누적 USD/KRW + 임계점 진행률 + provider/operation 분류',
    '월 사용량',
    '비용 카드 + 분류 표',
    'GET /api/admin/cost-monitor. 임계점 $100 (₩151,020)',
    g24='', iph16='', safari='●', chrome='●')
add('P6', 'A-026', '글로벌 임계점 알림 모달', 'Admin',
    '80%/100% 도달 시 자동 표시',
    'admin-web 진입',
    '알림 팝업 모달 + "확인 (N시간 후 재알림)"',
    '비용 임계점 도달',
    'snooze 24h (80%) / 2h (100%)',
    'GET /api/admin/critical-alerts. POST /alerts/{id}/dismiss',
    g24='', iph16='', safari='●', chrome='●')

# ─── P7 친구 초대 발급자 (mobile) ─────────────────────────────
add('P7', 'U-015', '친구 초대 코드 발급', 'USER',
    '홈 → "친구 초대"',
    '마이 또는 홈 → 초대',
    '초대 코드 발급 → SMS/이메일 공유 옵션',
    'kr_user',
    '코드 + 공유 시트',
    'POST /api/invitations')
add('P7', 'U-015', '내 초대 이력', 'USER',
    '내가 보낸 초대',
    '초대 화면',
    '내 발급 코드 목록 + 가입 여부',
    '초대 N건',
    '코드/상태 표',
    'GET /api/invitations')

# ─── P8 친구 초대로 가입한 신규 사용자 ─────────────────────────────
add('P8', 'U-003', '초대 코드 입력 가입', 'USER',
    '회원가입 화면 → 초대 코드',
    'register_screen',
    'invitation_code 입력 → 가입',
    'code=ABCD1234',
    '가입 완료 + 초대자 연결',
    'POST /api/auth/register with invitation_code. consume_invitation')
add('P8', 'U-015', '초대자 보상 확인', 'USER',
    '내 초대 이력',
    'P7 초대자가 확인',
    '친구 가입됨 표시',
    '초대 1건 used',
    '"가입 완료" 배지',
    '초대 보상 정책은 v1 미정 — 별도')

# ─── P9 만 14세 미만 + 보호자 ─────────────────────────────────────
add('P9', 'U-003', '14세 미만 가입 거부', 'USER',
    '회원가입 — birth_year=2020',
    'register_screen',
    'birth_year 2020 입력 → 가입 시도',
    'birth_year=2020',
    '400 + "보호자 동의 필요"',
    'register validate — age_group=minor / parent invite 안내')
add('P9', 'U-018', '보호자 초대 발급', 'USER',
    '보호자 이메일 입력',
    'parent_invite_screen',
    '보호자 이메일 → 동의 메일 발송',
    'parent@email',
    '동의 대기 안내',
    'POST /api/invitations/parent')
add('P9', 'U-018', '보호자 동의 → 가입 완료', 'USER',
    '보호자가 메일 링크 클릭',
    '동의 페이지',
    '동의 → 자녀 가입 자동 완료',
    'parent token',
    '자녀 계정 active',
    'GET /api/invitations/{code} + 가입 완료 트리거')

# ─── P10 소셜 로그인 ──────────────────────────────────────────
add('P10', 'U-002', '카카오 로그인', 'USER',
    '로그인 화면 → "카카오로 시작"',
    'login_screen',
    'Kakao OAuth → /api/auth/social',
    'Kakao test acc',
    '가입 + JWT 발급',
    'POST /api/auth/social provider=kakao. 비번 없음')
add('P10', 'U-002', '구글 로그인', 'USER',
    '로그인 화면 → "구글로 시작"',
    'login_screen',
    'Google OAuth → /api/auth/social',
    'Google test acc',
    '가입 + JWT 발급',
    'POST /api/auth/social provider=google')
add('P10', 'U-002', '애플 로그인', 'USER',
    '로그인 화면 → "Apple 로 시작"',
    'login_screen (iOS 만)',
    'Apple Sign In → /api/auth/social',
    'Apple test acc',
    '가입 + JWT 발급',
    'provider=apple. iOS 만 노출',
    g24='', iph16='●')
add('P10', 'U-016', '소셜 사용자 계정 삭제', 'USER',
    '설정 → 계정 삭제',
    '마이 → 설정',
    '비번 입력 단계 skip (소셜은 비번 없음)',
    'kakao 사용자',
    '즉시 삭제',
    'DELETE /api/auth/me. provider=email 만 비번 재확인')

# ─── P11 비밀번호 재설정 ─────────────────────────────────────────
add('P11', 'U-019', '비밀번호 찾기 요청', 'USER',
    '로그인 화면 → "비밀번호 찾기"',
    'forgot_password_screen',
    '이메일 입력 → 재설정 링크 메일 발송',
    'kr_user@pathwave.test',
    '메일 발송 안내',
    'POST /api/auth/forgot-password. rate limit 3/min')
add('P11', 'U-019', '재설정 — 새 비번 등록', 'USER',
    '메일 링크 클릭',
    '재설정 페이지 (앱 또는 web)',
    '새 비번 + 확인 → 저장',
    'reset_token',
    '재설정 완료, 새 비번으로 로그인',
    'POST /api/auth/reset-password. 토큰 만료 24h')

# ─── P12 푸시 토큰 등록 ──────────────────────────────────────────
add('P12', 'U-020', '푸시 권한 동의', 'USER',
    '앱 첫 실행 — OS 권한 다이얼로그',
    'splash_screen',
    'OS 권한 다이얼로그 → 허용',
    '신규 사용자',
    '권한 허용 + FCM 토큰 발급',
    'POST /api/push (token, platform=fcm|apns, language)')
add('P12', 'U-013', '매장 공지 푸시 수신', 'USER',
    '백그라운드 상태',
    '매장 공지 발송',
    '잠금 화면에 푸시 알림 노출',
    '매장 공지 1건',
    '알림 표시 (lang 별 본문)',
    '디바이스 lang ko 사용자는 ko, ja 사용자는 en')
add('P12', 'U-020', '푸시 권한 거부 시', 'USER',
    '권한 다이얼로그 → 거부',
    'splash_screen',
    'OS 권한 거부 → 인앱 알림만',
    '거부 사용자',
    '인앱 알림 센터에만 표시',
    'push_tokens row 미등록')

# ─── P13 1:1 문의 ──────────────────────────────────────────────
add('P13', 'U-022', '문의 작성', 'USER',
    '마이 → 고객지원 → 1:1 문의',
    'support_screen',
    '카테고리 선택 + 제목 + 본문 → 제출',
    '카테고리 "결제 문의"',
    'ticket id 발급',
    'POST /api/support/tickets')
add('P13', 'U-022', '슈퍼어드민 답변 수신', 'USER',
    '내 문의 목록',
    '마이 → 문의 이력',
    '답변 알림 + 메시지 확인',
    'ticket id',
    '관리자 답변 표시',
    'GET /api/support/tickets/me/{tid}')
add('P13', 'U-022', '추가 메시지 전송 (양방향)', 'USER',
    '문의 상세',
    '문의 상세 화면',
    '답글 입력 → 전송',
    '추가 메시지',
    '메시지 row +1',
    'POST /api/support/tickets/me/{tid}/messages')
add('P13', 'A-016', '슈퍼어드민 답변 작성', 'Admin',
    '/dashboard/support',
    '고객지원 → 문의 큐',
    '문의 선택 → 답변 작성 → 발송',
    'pending ticket',
    'status=replied',
    'POST /api/admin/support/tickets/{tid}/reply',
    g24='', iph16='', safari='●', chrome='●')

# ─── P14 알림 발송 승인 흐름 ───────────────────────────────────
add('P14', 'P-018', '사장 발송 요청', 'Provider',
    '/dashboard/notifications',
    '알림 발송 → 신규',
    '제목/본문/대상 입력 → 제출 (status=pending)',
    '매장 공지',
    '발송 대기 표시',
    'POST /api/facilities/{fid}/notifications',
    g24='', iph16='', safari='●', chrome='●')
add('P14', 'A-008', '슈퍼어드민 검토 → 승인', 'Admin',
    '/dashboard/notifications',
    '알림 검토',
    '발송 요청 행 → "승인" 또는 "거절"',
    'pending notification',
    'approved → dispatch',
    'POST /api/admin/notifications/{nid}/approve → /dispatch',
    g24='', iph16='', safari='●', chrome='●')
add('P14', 'A-008', '블록리스트 단어 등록', 'Admin',
    '/dashboard/notifications → 블록리스트',
    '검토 정책',
    '금지 단어 추가 → 향후 발송 자동 거부',
    '"광고", "스팸"',
    '블록리스트 +1',
    'POST /api/admin/notifications/blocklist',
    g24='', iph16='', safari='●', chrome='●')
add('P14', 'U-013', '사용자 측 수신', 'USER',
    '대상 사용자',
    '승인된 알림',
    '푸시 + 인앱 알림 둘 다 수신',
    '대상 user',
    '알림 표시',
    'P8c 푸시 자동 번역 lang 별')


# ════════════════════════════════════════════════════════════════════════════
#  스타일
# ════════════════════════════════════════════════════════════════════════════
HEADER_FILL = PatternFill('solid', start_color='2C2C38', end_color='2C2C38')
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
DEFAULT_FONT = Font(name='맑은 고딕', size=10)
WRAP = Alignment(wrap_text=True, vertical='top', horizontal='left')
CENTER = Alignment(horizontal='center', vertical='center')

THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PERSONA_COLORS = {
    'P1':  'FFE4E1', 'P2':  'E6F0FF', 'P3':  'E8F5E9', 'P4':  'C8E6C9',
    'P5':  'B2DFDB', 'P6':  'D1C4E9', 'P7':  'FFF3E0', 'P8':  'FFE0B2',
    'P9':  'FCE4EC', 'P10': 'F0F4C3', 'P11': 'B2EBF2', 'P12': 'DCEDC8',
    'P13': 'E1BEE7', 'P14': 'FFCCBC',
}


def build_cover(ws):
    ws.title = 'Cover'
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 80

    ws['B2'] = 'PathWave 페르소나 통합 테스트 시나리오'
    ws['B2'].font = Font(name='맑은 고딕', size=18, bold=True)
    ws['B3'] = f'작성일: {datetime.now().strftime("%Y-%m-%d")} · IA 와 같이 관리 (build_test_scenarios.py)'
    ws['B3'].font = Font(name='맑은 고딕', size=11, color='5C6170')

    # 페르소나
    ws['B5'] = '페르소나 14명'
    ws['B5'].font = Font(name='맑은 고딕', size=14, bold=True)
    for i, (pid, name, device, focus) in enumerate(PERSONAS, start=6):
        ws[f'B{i}'] = pid
        ws[f'B{i}'].font = Font(name='맑은 고딕', size=10, bold=True, color='2563EB')
        ws[f'C{i}'] = f'{name} · {device} · {focus}'
        ws[f'C{i}'].font = DEFAULT_FONT
        ws[f'C{i}'].alignment = WRAP

    base = 6 + len(PERSONAS)

    # 결과 기호
    ws[f'B{base + 2}'] = '결과 기호'
    ws[f'B{base + 2}'].font = Font(name='맑은 고딕', size=12, bold=True)
    symbols = [
        ('●',     '디바이스/브라우저 대상 — 테스트 수행'),
        ('(공백)', '디바이스/브라우저 비대상'),
        ('Pass',  '기대결과 일치'),
        ('Fail',  '기대결과 불일치 — 비고에 사유 + GitHub Issue 링크'),
        ('Skip',  '환경 미준비 (예: 실 비콘 / 실 키) — 다음 라운드'),
    ]
    for i, (s, d) in enumerate(symbols, start=base + 4):
        ws[f'B{i}'] = s
        ws[f'B{i}'].font = Font(name='맑은 고딕', size=10, bold=True, color='22C55E')
        ws[f'C{i}'] = d
        ws[f'C{i}'].font = DEFAULT_FONT

    # 컬럼 의미
    base2 = base + 4 + len(symbols) + 2
    ws[f'B{base2}'] = '컬럼 의미'
    ws[f'B{base2}'].font = Font(name='맑은 고딕', size=12, bold=True)
    col_notes = [
        ('화면ID',         'IA xlsx 의 화면 ID (U-001 / P-001 / A-001) 와 1:1 매칭. 신규 화면 추가 시 IA 동시 갱신.'),
        ('구분',           'USER (mobile) / Provider (provider-web) / Admin (admin-web)'),
        ('메뉴/화면',       '입력/선택/저장/조회 액션이 일어나는 정확한 위치'),
        ('가이드/진입',     '어떻게 그 화면에 진입하는지 (홈에서 N탭 등)'),
        ('테스트 데이터',    '계정/ID/세부 입력값 — R2 단계에서 테스트 계정 시드와 일치 시킴'),
        ('기대결과',        '코드 동작 기준 — API status/DB 변화/UI 표시'),
        ('디바이스 4종',    'Galaxy S24 / iPhone 16 Pro / Safari / Chrome — ● 표시된 곳만 실행'),
        ('Pass 여부',       '테스터 확인 후 채우는 컬럼 — Pass / Fail / Skip'),
        ('비고',           '결합내용 / 페르소나 ID / 디펜던시 / 이슈 링크'),
        ('테스터',          '확인자 이름 (개발팀 / QA / 사용자)'),
    ]
    for i, (k, v) in enumerate(col_notes, start=base2 + 2):
        ws[f'B{i}'] = k
        ws[f'B{i}'].font = Font(name='맑은 고딕', size=10, bold=True, color='2563EB')
        ws[f'C{i}'] = v
        ws[f'C{i}'].font = DEFAULT_FONT
        ws[f'C{i}'].alignment = WRAP

    # 관리 가이드
    base3 = base2 + 2 + len(col_notes) + 2
    ws[f'B{base3}'] = '지속 관리 가이드'
    ws[f'B{base3}'].font = Font(name='맑은 고딕', size=12, bold=True)
    guide = [
        ('1. 재생성',      '/Users/m5pro16/Desktop/pathwave/venv/bin/python docs/build_test_scenarios.py'),
        ('2. 신규 시나리오', 'docs/build_test_scenarios.py 의 ROWS 에 add(...) 추가 후 재생성'),
        ('3. IA 동기화',   '화면ID 가 IA xlsx 와 어긋나면 즉시 갱신. 신규 IA 행 추가 시 시나리오 후보 검토'),
        ('4. R2 테스터 분배', '페르소나별 시트(Persona-Index) 보고 테스터에 분배. Pass 컬럼은 실제 테스터가 직접 입력'),
        ('5. Fail 처리',    '비고 컬럼에 GitHub Issue 링크 + 재시도 일정. 재시도 후 다시 Pass 처리'),
        ('6. 결과 보고',    'Persona-Index 시트의 Pass / Fail / Skip 카운트로 진행률 파악'),
    ]
    for i, (k, v) in enumerate(guide, start=base3 + 2):
        ws[f'B{i}'] = k
        ws[f'B{i}'].font = Font(name='맑은 고딕', size=10, bold=True, color='22C55E')
        ws[f'C{i}'] = v
        ws[f'C{i}'].font = DEFAULT_FONT
        ws[f'C{i}'].alignment = WRAP


def build_scenarios(ws):
    ws.title = 'Scenarios'
    # 헤더
    for col, (label, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'C2'
    ws.row_dimensions[1].height = 40

    # 데이터
    for r_idx, row in enumerate(ROWS, start=2):
        # 비고 컬럼에서 persona 추출 (P1: 등)
        notes = row[15] or ''
        persona = notes.split(':', 1)[0].strip() if ':' in notes else (notes if notes.startswith('P') else '')
        fill_color = PERSONA_COLORS.get(persona, 'FFFFFF')
        fill = PatternFill('solid', start_color=fill_color, end_color=fill_color)

        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = DEFAULT_FONT
            c.border = BORDER
            if c_idx in (1,):                                # No 가운데
                c.alignment = CENTER
            elif c_idx in (11, 12, 13, 14, 15):              # 디바이스/Pass 가운데
                c.alignment = CENTER
            else:
                c.alignment = WRAP
            # No 컬럼만 색상 (페르소나 식별)
            if c_idx in (1, 2, 4):
                c.fill = fill
        ws.row_dimensions[r_idx].height = 50


def build_persona_index(ws):
    ws.title = 'Persona-Index'
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 14
    for col, label in enumerate(['ID', '페르소나', '시나리오 수', 'Pass', 'Fail', 'Skip', '미실행'], start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 12

    for i, (pid, name, _, _) in enumerate(PERSONAS, start=2):
        # 'P1:' / 'P10:' 콜론 매칭으로 정확히 카운트 (P1* 가 P10 도 매칭하던 버그 fix)
        prefix = f'{pid}:'
        count = sum(1 for r in ROWS if (r[15] or '').startswith(prefix))
        ws.cell(row=i, column=1, value=pid).font = Font(name='맑은 고딕', size=10, bold=True)
        ws.cell(row=i, column=2, value=name).font = DEFAULT_FONT
        ws.cell(row=i, column=3, value=count).alignment = CENTER
        # Pass/Fail/Skip — Scenarios 시트의 비고(P) + Pass(O) 컬럼 COUNTIFS
        ws.cell(row=i, column=4, value=f'=COUNTIFS(Scenarios!P:P,"{prefix}*",Scenarios!O:O,"Pass")').alignment = CENTER
        ws.cell(row=i, column=5, value=f'=COUNTIFS(Scenarios!P:P,"{prefix}*",Scenarios!O:O,"Fail")').alignment = CENTER
        ws.cell(row=i, column=6, value=f'=COUNTIFS(Scenarios!P:P,"{prefix}*",Scenarios!O:O,"Skip")').alignment = CENTER
        ws.cell(row=i, column=7,
                value=f'=C{i}-D{i}-E{i}-F{i}').alignment = CENTER

    total_row = 2 + len(PERSONAS)
    ws.cell(row=total_row, column=2, value='합계').font = Font(name='맑은 고딕', size=11, bold=True)
    for col, letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (7, 'G')]:
        ws.cell(row=total_row, column=col,
                value=f'=SUM({letter}2:{letter}{total_row - 1})')
        ws.cell(row=total_row, column=col).font = Font(name='맑은 고딕', size=11, bold=True)
        ws.cell(row=total_row, column=col).alignment = CENTER


# ════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    cover = wb.active
    build_cover(cover)
    scenarios = wb.create_sheet('Scenarios')
    build_scenarios(scenarios)
    index_ws = wb.create_sheet('Persona-Index')
    build_persona_index(index_ws)

    out = '/Users/m5pro16/Desktop/pathwave/docs/pathwave_persona_test_scenarios_2026-05-26.xlsx'
    wb.save(out)
    print(f'wrote {out}')
    print(f'total scenarios: {len(ROWS)} · personas: {len(PERSONAS)}')


if __name__ == '__main__':
    main()
