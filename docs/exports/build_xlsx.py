"""PathWave — 개발 체크리스트 + 서비스 신청·운영비용 엑셀 2개 생성 (v5).

v5 수정사항 (2026-05-23):
  - "BE" → "Pathwave 서비스" 명칭 통일
  - 시나리오: MAU 기반 → 매장 수 기반 (매장당 매출 ₩10,000 직결)
  - 매장당 평균 사용자(MAU) 입력셀 분리 — 자동 계산
  - 서버 옵션: Mac mini 자체구축(0원) + 가비아 호스팅(5K) — 클라우드 제외
  - 알림톡 → 옵션화 (자체 앱 푸시·이메일·웹 푸시로 대체 가능)
  - SMS 인증 → 옵션화 (이메일 인증으로 대체 가능)
  - 자동화 도구 → Mac mini 자체구축 (외부 SaaS 제거/옵션화)
  - 던스(D-U-N-S) 번호 신청 추가 (무료, Apple Developer 법인 가입 선행)
  - 이메일 수량별 SendGrid tier 명시
  - 기능 sub-feature 추가 세분화
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

USD_KRW = 1500

FONT = '맑은 고딕'
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SUBHEADER_FILL = PatternFill('solid', start_color='D9E1F2')
INPUT_FONT  = Font(name=FONT, color='0000FF')
LINK_FONT   = Font(name=FONT, color='008000')
ASSUME_FILL = PatternFill('solid', start_color='FFFFCC')
DONE_FILL    = PatternFill('solid', start_color='C6EFCE')
PARTIAL_FILL = PatternFill('solid', start_color='FFEB9C')
ACTIVE_FILL  = PatternFill('solid', start_color='BDD7EE')
TODO_FILL    = PatternFill('solid', start_color='F2F2F2')
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')

KRW_FMT = '₩#,##0;(₩#,##0);-'
PCT_FMT = '0.0%;-0.0%;-'
DATE_FMT = 'yyyy-mm-dd'

ACTOR_COLOR = {
    '사용자':         'E2EFDA',
    '시설관리자':     'FFF2CC',
    '슈퍼어드민':     'FCE4D6',
    'Pathwave 서비스': 'DEEBF7',
    '인프라':         'EAEAEA',
}


def force_workbook_font(wb, font_name=None):
    fn = font_name or FONT
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                f = cell.font
                cell.font = Font(name=fn, size=f.size, bold=f.bold, italic=f.italic,
                                 color=f.color, underline=f.underline, strike=f.strike)


def apply_table(ws, header_row, last_row, last_col, freeze=True):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for r in range(header_row + 1, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if not cell.alignment.wrap_text:
                cell.alignment = WRAP
    if freeze:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


# ═══════════════════════════════════════════════════════════════════════════
# FILE 1 — 개발 체크리스트
# ═══════════════════════════════════════════════════════════════════════════
def build_dev_checklist():
    wb = Workbook()
    wb.remove(wb.active)
    today = '2026-05-23'

    ws = wb.create_sheet('기능별 전체 현황')
    ws['A1'] = 'PathWave 기능별 전체 현황 — 주체별 sub-feature × 이전/Phase 1/남은'
    ws['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:G1')

    headers = ['주체', '도메인/기능', '단계', 'sub-feature', '내용', '상태', '비고']
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    items = [
        # ════════ 인증/회원가입 ════════
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '이메일 인증', '/api/auth send-code·verify-code + email_codes 테이블 + 5분 만료', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '이메일 회원가입', '/api/auth/register + bcrypt + 약관 동의 record_consents', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '이메일 로그인', '/api/auth/login + JWT 발급(access+refresh)', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '5종 소셜 로그인 토큰 검증', 'Google·Apple·Facebook·Kakao·Naver token verify + 사용자 매핑', '✅', '실 키는 카카오·네이버 검수 후'),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '동의·약관 8 KIND', 'policies(terms/privacy/location/marketing/push/camera/storage/third_party/age14) 등록·버전관리', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '미성년자 보호 + 부모 초대', 'parent_invite 코드 발급/검증 + age14 동의 흐름', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '회원 탈퇴 + 14일 그레이스', '/api/auth/delete-account + soft delete + deleted_at + 재가입 차단', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '모바일 비번 재설정 API', '/api/auth/forgot-password·reset-password (users 테이블)', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', '이전 완료', '시설 계정 가입·로그인', '/api/facility send-code·register·login + facility_accounts + 운영자 승인 대기', '✅', ''),
        ('Pathwave 서비스', '인증/회원가입', 'Phase 1 ✅', 'P4 — 시설 비번 재설정 API 신규', '/api/facility/forgot-password·reset-password 신규(email_codes 재활용)', '✅', ''),
        ('사용자', '인증/회원가입', '이전 완료', 'mobile 가입 5단계 흐름', 'consent → 본인인증 → 이메일/소셜 → 약관 → 완료', '✅', ''),
        ('사용자', '인증/회원가입', '이전 완료', 'mobile login_screen', '이메일 로그인 + 5종 소셜 버튼 + forgot password 링크', '✅', ''),
        ('사용자', '인증/회원가입', '이전 완료', 'mobile forgot_password_screen', '2단계 — 이메일 입력 → 코드+새비번', '✅', ''),
        ('사용자', '인증/회원가입', '이전 완료', 'mobile delete_account_screen', '탈퇴 안내·확인 + 14일 그레이스 안내', '✅', ''),
        ('시설관리자', '인증/회원가입', '이전 완료', 'provider 회원가입 폼', '실 폼 + 매장 정보 + 소셜 로그인', '✅', ''),
        ('시설관리자', '인증/회원가입', '이전 완료', 'provider Login.jsx', '실 폼 존재했으나 dev 우회로 가려졌었음', '✅', ''),
        ('시설관리자', '인증/회원가입', 'Phase 1 ✅', 'P4 — dev 우회 env 게이트', 'DEV_AUTO_LOGIN = import.meta.env.DEV (출시빌드 비활성)', '✅', ''),
        ('시설관리자', '인증/회원가입', 'Phase 1 ✅', 'P4 — Login 자동토큰 제거', 'useEffect 자동 mock 토큰 발급 제거 → 실 로그인 폼 노출', '✅', ''),
        ('시설관리자', '인증/회원가입', 'Phase 1 ✅', 'P4 — Signup 게스트 게이트', '"로그인 없이 이용" 버튼 dev 환경에서만 노출', '✅', ''),
        ('시설관리자', '인증/회원가입', 'Phase 1 ✅', 'P4 — ForgotPassword 신규', 'provider /forgot-password 라우트 + 2단계 UI + AuthService 메서드 추가', '✅', ''),
        ('슈퍼어드민', '인증/회원가입', '이전 완료', 'admin Login', 'admin-web Login + RequireAuth (DevPreviewBar는 VITE_PREVIEW_MODE 게이트됨)', '✅', ''),

        # ════════ 매장 정보 ════════
        ('Pathwave 서비스', '매장 정보', '이전 완료', '매장 CRUD API', '/api/facilities POST/GET/PATCH/DELETE + owner 권한', '✅', ''),
        ('Pathwave 서비스', '매장 정보', '이전 완료', '매장 이미지 갤러리', '/api/facilities/<id>/images CRUD', '✅', ''),
        ('Pathwave 서비스', '매장 정보', '이전 완료', '다국어 번역 캐시', 'facilities/translations join — 12언어', '✅', ''),
        ('Pathwave 서비스', '매장 정보', '이전 완료', '비콘 claim·list API', '/api/store/<id>/beacons + /claim-beacon (Phase C)', '✅', ''),
        ('Pathwave 서비스', '매장 정보', '이전 완료', 'company_info API (Phase M)', '슈퍼어드민 입력 GET /api/company-info', '✅', ''),
        ('사용자', '매장 정보', '이전 완료', 'mobile facility_screen', '이미지/영업시간/주소/카테고리/채팅 진입 + 즐겨찾기', '✅', ''),
        ('시설관리자', '매장 정보', '이전 완료', 'provider StoreInfo UI 골격', '편집/저장/지도(leaflet)/이미지/카테고리/영업시간/휴무/비콘 claim — 데이터는 mock', '✅', ''),
        ('시설관리자', '매장 정보', 'Phase 1 ✅', 'P5 — StoreInfo 마운트 로드', 'StoreService.list() → 1계정1매장 → setStore + fid', '✅', ''),
        ('시설관리자', '매장 정보', 'Phase 1 ✅', 'P5 — handleSave PATCH 연동', 'StoreService.update(fid, payload) — business_hours 양방향 매핑', '✅', ''),
        ('시설관리자', '매장 정보', 'Phase 1 ✅', 'P5 — 비콘 fid 실연동', 'MOCK_FID="demo" 제거 → 실 fid로 fetchBeacons·claimBeacon', '✅', ''),
        ('시설관리자', '매장 정보', 'Phase 1 ✅', 'P5 — Settings 푸터 PwFooter', '하드코딩 시원컴퍼니/서초구 → CompanyInfoService 실 법인정보', '✅', ''),
        ('시설관리자', '매장 정보', 'Phase 1 ✅', 'P5 — Facilities dead code 삭제', '1계정1매장 위반 + 라우트 미연결 파일 + import 제거', '✅', ''),
        ('슈퍼어드민', '매장 정보', '이전 완료', 'admin Stores 관리', '매장 관리 + 비콘 매핑 + 번역 등록', '✅', ''),
        ('Pathwave 서비스', '매장 정보', '남은', 'P8c — 스키마 갭', 'facilities 테이블 category/holidays/detailAddress 컬럼 추가 또는 UI 정리', '⬜', '별도 작업 칩'),

        # ════════ 결제·구독 ════════
        ('Pathwave 서비스', '결제·구독', '이전 완료', '카드 등록·삭제·목록 API', '/api/billing/cards POST/GET/DELETE — last4·PG key 토큰화', '✅', ''),
        ('Pathwave 서비스', '결제·구독', '이전 완료', '구독 생성·해지·연장 API', '/api/billing/subscriptions + cancel + extend', '✅', ''),
        ('Pathwave 서비스', '결제·구독', '이전 완료', '결제 내역·환불 API', '/api/billing/payments + admin refund', '✅', ''),
        ('Pathwave 서비스', '결제·구독', '이전 완료', '영수증 이메일 + PG 추상화', '/api/billing/receipt-email + PG provider(sim/toss)', '✅', '키 활성화 시 toss 전환'),
        ('시설관리자', '결제·구독', '이전 완료', 'provider PaymentManagement UI 골격', '다단계 결제 흐름/카드 입력 모달/내역/구독 카드 — 데이터는 mock + localStorage 평문저장', '✅', ''),
        ('시설관리자', '결제·구독', 'Phase 1 ✅', 'P7 — BillingService 신규', '7종 메서드 래퍼 (cards/subs/payments/receipt-email) + PCI 주석', '✅', ''),
        ('시설관리자', '결제·구독', 'Phase 1 ✅', 'P7 — 카드 localStorage 평문저장 제거 (C6)', 'pathwave_payment_card + audit 큐 완전 제거 → brand+last4 만 전송', '✅', 'PCI 준수'),
        ('시설관리자', '결제·구독', 'Phase 1 ✅', 'P7 — 결제 흐름 통합', 'PaymentManagement·Subscriptions·ServiceRequest 동일 백엔드 통합', '✅', ''),
        ('슈퍼어드민', '결제·구독', '이전 완료', 'admin Payments + Subscriptions + 환불', 'adminApi.listPayments/listSubscriptions/refundPayment 실연동 + 환불 모달', '✅', ''),
        ('Pathwave 서비스', '결제·구독', '남은', '토스 키 활성화 + 실결제', '심사 통과 후 PG_PROVIDER=toss + 실 결제 end-to-end', '⬜', '외부 1~2주'),

        # ════════ 쿠폰 ════════
        ('Pathwave 서비스', '쿠폰', '이전 완료', '쿠폰 발행/사용/만료 API', '/api/coupons issue·use·expire + 정책 + 통계', '✅', ''),
        ('사용자', '쿠폰', '이전 완료', 'mobile coupons_screen', '활성/사용/만료 탭 + 사용 다이얼로그', '✅', '사용 silent error 잔존'),
        ('사용자', '쿠폰', 'Phase 1 — 남은', 'P9 — mobile silent error 수정', '사용 실패 시 사용자 피드백 + 목록 갱신', '⬜', ''),
        ('사용자', '쿠폰', 'Phase 1 — 남은', 'P22 — 회원 QR 발급(URL 토큰)', '마이페이지 회원 QR + 토큰 만료/재발급', '⬜', ''),
        ('시설관리자', '쿠폰', '이전 완료', 'provider Coupons UI 골격', '쿠폰 목록 UI 완성, 데이터 mock', '✅', ''),
        ('시설관리자', '쿠폰', '이전 완료', 'provider CouponForm UI 골격', '쿠폰 생성·통계 UI', '✅', ''),
        ('시설관리자', '쿠폰', 'Phase 1 — 남은', 'P9 — provider Coupons 실연동', 'mock → /api/coupons 실연동', '⬜', ''),
        ('시설관리자', '쿠폰', 'Phase 1 — 남은', 'P22 — provider QR 스캔/적립·사용', '회원 QR 스캔 + 코드 입력 → 적립/사용', '⬜', ''),
        ('슈퍼어드민', '쿠폰', '이전 완료', 'admin 쿠폰 통계 API', 'admin coupon stats', '✅', 'CouponStats placeholder → P11'),
        ('슈퍼어드민', '쿠폰', 'Phase 1 — 남은', 'P11 — CouponStats 실연동', 'admin placeholder 해소', '⬜', ''),

        # ════════ 스탬프 ════════
        ('Pathwave 서비스', '스탬프', '이전 완료', '스탬프 적립 API + 정책', '/api/stamps accrue/list + 시설별 정책', '✅', ''),
        ('Pathwave 서비스', '스탬프', '이전 완료', '비콘 자동 적립', 'BLE 비콘 감지 → 자동 적립', '✅', ''),
        ('사용자', '스탬프', '이전 완료', 'mobile stamps_screen', '시설별 카드/진행률/보상 표시', '✅', ''),
        ('시설관리자', '스탬프', '이전 완료', 'provider Stamps UI 골격', '정책·통계 UI', '✅', ''),
        ('시설관리자', '스탬프', '이전 완료', 'provider StampForm UI 골격', '스탬프 정책 폼', '✅', ''),
        ('시설관리자', '스탬프', 'Phase 1 — 남은', 'P9 — provider Stamps 실연동', 'mock → /api/stamps 실연동', '⬜', ''),
        ('시설관리자', '스탬프', 'Phase 1 — 남은', 'P22 — 수동 적립(staff)', '회원 QR 스캔 → 수동 적립 (auto/staff 모드)', '⬜', ''),

        # ════════ 채팅 ════════
        ('Pathwave 서비스', '채팅', '이전 완료', '채팅방·메시지 API', '/api/chat rooms·messages + 페이지네이션', '✅', ''),
        ('Pathwave 서비스', '채팅', '이전 완료', 'SSE 실시간 스트림', 'GET /api/chat/rooms/<id>/stream + 5분 의도 종료 + after_id 지원', '✅', ''),
        ('Pathwave 서비스', '채팅', '이전 완료', '읽음 처리 + chat_blocks', 'POST /read + is_blocked 가드', '✅', ''),
        ('Pathwave 서비스', '채팅', '이전 완료', 'abuse_reports', '/api/abuse-reports + /api/admin/abuse-reports CRUD + PATCH', '✅', ''),
        ('Pathwave 서비스', '채팅', 'Phase 1 ✅', 'P8 — /api/chat/reports 신규', 'abuse_reports → ChatMonitor 형태 변환, super_admin 가드', '✅', ''),
        ('사용자', '채팅', '이전 완료', 'mobile chat_list_screen', '내 채팅방 목록 + 미읽음 배지', '✅', ''),
        ('사용자', '채팅', '이전 완료', 'mobile chat_detail_screen', 'SSE 실시간 + 신고/차단 메뉴 + 가이드라인 모달', '✅', '재연결 미지원이었음'),
        ('사용자', '채팅', 'Phase 1 ✅', 'P8 — mobile SSE 재연결', 'StreamController + 지수 backoff(1→30s) + ?after_id= 누락 보충', '✅', '⚠ app.py 재시작 후 활성'),
        ('사용자', '채팅', '남은 — P8b', 'mobile 번역 표시 UI', 'translation 필드 실데이터 렌더', '⬜', ''),
        ('시설관리자', '채팅', '이전 완료', 'provider ChatService 메서드', 'openRoom/listRooms/listMessages/sendMessage/markRead/subscribe(SSE) — 작성됨, 호출 안 됨', '✅', ''),
        ('시설관리자', '채팅', '이전 완료', 'provider CustomerChat UI 골격', '대화창 UI 완성, DUMMY_CHATS 7개 더미', '✅', ''),
        ('시설관리자', '채팅', 'Phase 1 ✅', 'P8 — CustomerChat 실연동', 'DUMMY 제거 + ChatService 5종 호출 + 30초 폴링 + 낙관적 전송 + SSE 정리', '✅', ''),
        ('시설관리자', '채팅', '남은 — P8b', 'provider 번역 표시', 'translation 필드 실데이터 노출', '⬜', ''),
        ('슈퍼어드민', '채팅', '이전 완료', 'admin abuse·chat_blocks 처리', 'abuse PATCH 상태 변경 + chat_blocks 관리', '✅', 'PR #142'),
        ('슈퍼어드민', '채팅', '이전 완료', 'admin ChatMonitor UI 골격', '신고 큐 테이블 UI 완성 (엔드포인트만 없었음)', '✅', ''),
        ('Pathwave 서비스', '채팅', '남은 — P8b', '번역 캐시 스키마 신규', 'chat_message_translations (message_id, lang → translated)', '⬜', ''),
        ('Pathwave 서비스', '채팅', '남은 — P8b', 'list_messages·SSE 번역 머지', '?lang=xx 응답에 translated_text 포함 + translator.py 호출', '⬜', '키 활성화 후'),

        # ════════ 알림 (자체 푸시 + 이메일 중심) ════════
        ('Pathwave 서비스', '알림', '이전 완료', '알림 inbox·공지 API', '/api/notifications inbox·announcements·read', '✅', ''),
        ('Pathwave 서비스', '알림', '이전 완료', '알림 선호 카테고리', 'notification_preferences API', '✅', ''),
        ('Pathwave 서비스', '알림', '이전 완료', 'iOS APNs 자체 (PR #50)', 'APNs HTTP/2 + JWT ES256 — 자체구축, 인증서/키만 필요', '✅', ''),
        ('Pathwave 서비스', '알림', '이전 완료', 'Android FCM', 'PR #44 — Google Play Services 의존 (무료 무제한)', '✅', '⚠ Android는 FCM 외 사실상 대안 없음'),
        ('Pathwave 서비스', '알림', '남은 (자체구축)', '웹 푸시(VAPID)', '시설관리자 웹 푸시 자체구축 — Service Worker + Web Push API', '⬜', '시설관리자 알림용'),
        ('사용자', '알림', '이전 완료', 'mobile notifications_screen', '인박스/공지 탭', '✅', '라우팅 보강 필요'),
        ('사용자', '알림', '이전 완료', 'mobile 권한 동의 다이얼로그', 'notification_permission_dialog', '✅', ''),
        ('사용자', '알림', 'Phase 1 — 남은', 'P10 — mobile 알림 라우팅', '알림 종류별 화면 이동 (coupon→쿠폰함, chat→채팅)', '⬜', ''),
        ('시설관리자', '알림', '이전 완료', 'provider Notifications UI 골격', '받은 알림 + 시스템 공지 UI, mockInbox 14건', '✅', ''),
        ('시설관리자', '알림', 'Phase 1 — 남은', 'P10 — provider Notifications 실연동', 'mockInbox → /api/notifications 실연동', '⬜', ''),
        ('시설관리자', '알림', 'Phase 1 — 남은', 'P10 — provider 웹 푸시 통합', '브라우저 Web Push (자체구축, VAPID)', '⬜', ''),
        ('슈퍼어드민', '알림', '이전 완료', 'admin 공지 CRUD + 발송', 'Announcements CRUD + 푸시 발송 통합', '✅', ''),

        # ════════ WiFi (1회 연결) ════════
        ('Pathwave 서비스', 'WiFi (1회 연결)', '이전 완료', 'wifi_profiles 기본', 'wifi_profiles 테이블 + CRUD API', '✅', ''),
        ('Pathwave 서비스', 'WiFi (1회 연결)', '이전 완료', 'BLE 핸드셰이크 API', '/api/beacon/wifi handshake (LIMIT 1)', '✅', '묶음 반환 확장 P15'),
        ('사용자', 'WiFi (1회 연결)', '이전 완료', 'mobile iOS native plugin', 'NEHotspotConfiguration 자동 가입 요청 (PR #49)', '✅', ''),
        ('사용자', 'WiFi (1회 연결)', '이전 완료', 'mobile Android native plugin', 'WifiNetworkSuggestion 자동 가입', '✅', ''),
        ('사용자', 'WiFi (1회 연결)', '이전 완료', 'mobile WifiConnectScreen', 'BLE 핸드셰이크 → OS 자동 가입 트리거', '✅', ''),
        ('시설관리자', 'WiFi (1회 연결)', '이전 완료', 'provider WifiSettings UI', '등록·수정·비활성 UI 완성', '✅', ''),
        ('시설관리자', 'WiFi (1회 연결)', 'Phase 1 ✅', 'P6 — runOcrMock 제거', 'WifiSettings·ServiceRequest 가짜 OCR 자동입력 삭제', '✅', '심의 reject 방지'),
        ('시설관리자', 'WiFi (1회 연결)', 'Phase 1 ✅', 'P6 — 사진 첨부 재정의', '사진 = 입력 확인용 참고 첨부 (정직한 안내문구)', '✅', ''),

        # ════════ WiFi 로밍 (B 풀스코프) ════════
        ('Pathwave 서비스', 'WiFi 로밍', 'Phase 1 — 남은', 'P14 — 데이터 모델 재설계', 'wifi_profiles 확장 + beacon_wifi·units·wifi_access_grant·devices 신규', '⬜', ''),
        ('Pathwave 서비스', 'WiFi 로밍', 'Phase 1 — 남은', 'P14 — beacons.role 컬럼 추가', 'wifi / cashier 구분', '⬜', ''),
        ('Pathwave 서비스', 'WiFi 로밍', 'Phase 1 — 남은', 'P15 — handshake 묶음 반환', 'LIMIT 1 제거 → 매장 WiFi 묶음', '⬜', ''),
        ('시설관리자', 'WiFi 로밍', 'Phase 1 — 남은', 'P15 — WifiSettings 실연동 + 비콘 role UI', '실 API 연동 + 비콘 role 선택', '⬜', ''),
        ('슈퍼어드민', 'WiFi 로밍', 'Phase 1 — 남은', 'P15 — admin WiFi 등록 신규', '슈퍼어드민 WiFi 등록 화면 신규', '⬜', 'C10'),
        ('사용자', 'WiFi 로밍', 'Phase 1 — 남은', 'P16 — 비콘→WiFi 묶음 fetch + 캐시', '클라이언트 캐시 전략', '⬜', ''),
        ('사용자', 'WiFi 로밍', 'Phase 1 — 남은', 'P16 — BSSID 검증', '연결 후 BSSID 매칭 확인', '⬜', ''),
        ('사용자', 'WiFi 로밍', 'Phase 1 — 남은', 'P16 — "WiFi 변경됨" 흐름 + 손님 자동/승인', '비콘 이동 감지 → 자동 전환', '⬜', ''),
        ('사용자', 'WiFi 로밍', 'Phase 1 — 남은', 'P17 — .mobileconfig 다건 설치', '.mobileconfig 생성·다건 설치 (서명은 인증서 후)', '⬜', ''),
        ('Pathwave 서비스', 'WiFi 로밍', 'Phase 1 — 남은', 'P18 — credential_mode managed', '비번 교체 리마인드 + 인가 손님 자동 전파 + 알림 연동', '⬜', 'v1 flag 비공개'),
        ('슈퍼어드민', 'WiFi 로밍', 'Phase 1 — 남은', 'P19 — units/grant 관리 UI', '호실·자리 시간제 권한 UI (admin/provider)', '⬜', 'v1 flag 비공개'),

        # ════════ 비콘 ════════
        ('Pathwave 서비스', '비콘', '이전 완료', 'beacons lifecycle', 'inventory → active → inactive/lost 상태 전이 API', '✅', ''),
        ('Pathwave 서비스', '비콘', '이전 완료', '비콘 claim API', '시설 가입 후 비콘 SN 클레임', '✅', ''),
        ('슈퍼어드민', '비콘', '이전 완료', 'admin Beacons 인벤토리', 'CSV 입고 + claim 모니터 + 배터리/펌웨어 표시', '✅', ''),
        ('시설관리자', '비콘', '이전 완료', 'provider 비콘 claim UI', 'StoreInfo 내 claim + listBeacons', '✅', ''),

        # ════════ i18n (12언어) ════════
        ('Pathwave 서비스', 'i18n', '이전 완료', 'translations 테이블 + API', 'GET /api/i18n/{lang}', '✅', ''),
        ('Pathwave 서비스', 'i18n', '이전 완료', 'DeepL 통합 모듈', 'translator.py — DeepL/Google/Stub provider', '✅', '키 활성화 후 batch'),
        ('슈퍼어드민', 'i18n', '이전 완료', 'admin i18n CRUD', '키별 다국어 입력·자동번역 트리거', '✅', ''),
        ('사용자', 'i18n', '이전 완료', 'mobile supportedLocales 7개 → 12개', 'P2 이전 7개, 이후 12개로 확장', '✅', ''),
        ('사용자', 'i18n', 'Phase 1 ✅', 'P2 — mobile I18nService 통일', '12 supportedLocales + 전 화면 t() 전환', '✅', ''),
        ('Pathwave 서비스', 'i18n', 'Phase 1 ✅', 'P2 — ko 시드 550 + DeepL 스크립트', 'translations DB ko 203→550 시드 + translate_i18n_deepl.py', '✅', ''),
        ('Pathwave 서비스', 'i18n', '남은', 'DeepL 11언어 batch 실행', '키 활성화 후 22언어 × 549건 자동번역', '⬜', 'DeepL 키 후'),

        # ════════ 정책/약관 ════════
        ('Pathwave 서비스', '정책/약관', '이전 완료', 'policies API + 8 KIND', '버전 관리 + record_consents 감사 로그', '✅', ''),
        ('사용자', '정책/약관', '이전 완료', 'mobile policy_view_screen', '약관 본문 뷰어', '✅', ''),
        ('시설관리자', '정책/약관', '이전 완료', 'provider policy_view', '약관 본문 노출', '✅', ''),
        ('슈퍼어드민', '정책/약관', '이전 완료', 'admin Policies CRUD', '등록·버전 발행', '✅', ''),
        ('Pathwave 서비스', '정책/약관', 'Phase 1 — 남은', 'P13 — 환불·청소년·쿠키 3종 추가', 'BE policy KIND 추가', '⬜', 'C14'),
        ('시설관리자', '정책/약관', 'Phase 1 — 남은', 'P13 — provider 노출', '신규 3종 노출 + 동의 흐름', '⬜', ''),
        ('사용자', '정책/약관', 'Phase 1 — 남은', 'P13 — policy_view 언어 정합', '?lang=ko 강제 → 사용자 언어로', '⬜', ''),

        # ════════ 디자인 시스템 ════════
        ('사용자', '디자인 시스템', '이전 완료', 'mobile PwTheme + 보라', 'PwTheme 다크 + Pw* 위젯 9종', '✅', ''),
        ('사용자', '디자인 시스템', 'Phase 1 ✅', 'P1 — mobile 단일화 + Pw* 4종 추가', '테마 단일화 + 시스템 폰트 + Radio·Checkbox·Dropdown·Chip', '✅', ''),
        ('시설관리자', '디자인 시스템', '이전 완료', 'provider 다크 + 녹색', 'PR #66~#87 통합', '✅', ''),
        ('시설관리자', '디자인 시스템', 'Phase 1 ✅', 'P3 — provider 공용 모달', 'alert/confirm 16곳 → useDialog()', '✅', ''),
        ('슈퍼어드민', '디자인 시스템', '이전 완료', 'admin 다크 + 블루', 'PR #65', '✅', ''),
        ('슈퍼어드민', '디자인 시스템', 'Phase 1 ✅', 'P3 — admin 공용 모달', 'alert/confirm 17곳 → useDialog()', '✅', ''),

        # ════════ 슈퍼어드민 (대시보드·통계) ════════
        ('슈퍼어드민', '대시보드·통계', '이전 완료', 'admin 7페이지 베이스라인', 'Login + Dashboard + Beacons + Approvals + Battery + Announcements + Stores', '✅', 'PR #36~#38'),
        ('슈퍼어드민', '대시보드·통계', '이전 완료', 'admin Payments·Subscriptions·환불', '결제/구독/환불 — 실 API 연동', '✅', 'PR #39'),
        ('슈퍼어드민', '대시보드·통계', '이전 완료', 'admin Policies·Translations·Audit·Reports', '약관·번역·감사·신고 관리', '✅', ''),
        ('슈퍼어드민', '대시보드·통계', '이전 완료', 'admin 채팅 신고/차단 처리', 'abuse + chat_blocks (PR #142)', '✅', ''),
        ('슈퍼어드민', '대시보드·통계', 'Phase 1 — 남은', 'P11 — Dashboard 가짜데이터 제거', '실 통계 API 연동', '⬜', ''),
        ('슈퍼어드민', '대시보드·통계', 'Phase 1 — 남은', 'P11 — StaffMonitor 실연동', '시설별 직원 모니터링', '⬜', ''),
        ('슈퍼어드민', '대시보드·통계', 'Phase 1 — 남은', 'P11 — CouponStats placeholder 해소', '실 데이터 표시', '⬜', ''),
        ('슈퍼어드민', '대시보드·통계', 'Phase 1 — 남은', 'P20 — app-versions UI 신규', 'BE 완성된 routes/version.py에 admin UI 연결', '⬜', 'C11'),

        # ════════ 직원 관리 ════════
        ('Pathwave 서비스', '직원 관리', '이전 완료', 'staff API', '초대/권한/관리 백엔드', '✅', ''),
        ('시설관리자', '직원 관리', '이전 완료', 'provider StaffManagement UI 골격', '목록·초대·권한 UI, 데이터 mock', '✅', ''),
        ('시설관리자', '직원 관리', 'Phase 1 — 남은', 'P11 — provider StaffManagement 실연동', 'mock → 실 staff API', '⬜', ''),

        # ════════ mobile 심의 직격 ════════
        ('사용자', '심의 직격', '이전 완료', 'consent 단계', '8 KIND 동의 + 미성년자 + 약관 본문', '✅', 'placeholder 노출 버그'),
        ('사용자', '심의 직격', '이전 완료', 'settings 화면', '계정·알림·약관·고객지원·탈퇴', '✅', 'dev API URL 노출 + 버전 하드코딩'),
        ('사용자', '심의 직격', 'Phase 1 — 남은', 'P12 — consent placeholder 제거', 'placeholder 본문 노출 수정', '⬜', '심의 reject 방지'),
        ('사용자', '심의 직격', 'Phase 1 — 남은', 'P12 — 동의 로드 실패 복구', '네트워크 오류 시 재시도/안내', '⬜', ''),
        ('사용자', '심의 직격', 'Phase 1 — 남은', 'P12 — settings dev 정보 제거', 'API URL 노출 제거', '⬜', ''),
        ('사용자', '심의 직격', 'Phase 1 — 남은', 'P12 — 앱 버전 동적화', 'package_info_plus 도입 — 하드코딩 제거', '⬜', 'C9'),

        # ════════ 앱 버전관리 ════════
        ('Pathwave 서비스', '앱 버전관리', '이전 완료', 'BE routes/version.py + app_versions', 'admin GET/PUT + mobile GET /api/version', '✅', 'admin UI만 필요'),
        ('슈퍼어드민', '앱 버전관리', 'Phase 1 — 남은', 'P20 — admin app-versions UI', '버전 등록·강제/권장 토글·OS별', '⬜', ''),
        ('사용자', '앱 버전관리', 'Phase 1 — 남은', 'P20 — mobile 최소버전 게이트', 'splash 버전 체크 → 강제 업데이트 모달', '⬜', '일부 처리 존재'),

        # ════════ 심의 메타 자산 ════════
        ('인프라', '심의 메타', 'Phase 1 — 남은', 'P21 — iOS PrivacyInfo.xcprivacy', 'iOS 17.4+ 필수 — 데이터 수집/추적 명시', '⬜', '심의 reject 방지'),
        ('인프라', '심의 메타', 'Phase 1 — 남은', 'P21 — Bundle ID 확정', 'co.triggersoft.pathwave 등', '⬜', ''),
        ('인프라', '심의 메타', 'Phase 1 — 남은', 'P21 — Android Photo Picker', 'Android 13+ 신규 권한 모델', '⬜', ''),
        ('인프라', '심의 메타', 'Phase 1 — 남은', 'P21 — 계정삭제 웹 URL', 'Google Play 정책 — 외부 웹 URL', '⬜', ''),

        # ════════ 회사 정보 (footer) ════════
        ('사용자', '회사 정보', '이전 완료', 'mobile PwFooter', 'CompanyInfoService 실연동', '✅', ''),
        ('시설관리자', '회사 정보', '이전 완료', 'provider PwFooter', 'CompanyInfoService 실연동', '✅', ''),
        ('슈퍼어드민', '회사 정보', '남은', '실 법인정보 입력', '슈퍼어드민이 트리거소프트 실 정보 admin에서 입력', '⬜', '법인 등기 완료'),

        # ════════ 운영 인프라 ════════
        ('인프라', '운영 인프라', '이전 완료', 'BE 보안 강화', 'SECRET_KEY/AES_KEY ENV + CORS + rate-limit', '✅', 'PR #35'),
        ('인프라', '운영 인프라', '이전 완료', 'BE provider 추상화', 'PG(sim/toss) + Email + Push 추상화', '✅', '키 활성화 시 전환'),
        ('인프라', '운영 인프라', '이전 완료', 'BE gunicorn/wsgi + Sentry + Postgres', '프로덕션 배포 골격 + DATABASE_URL 분기', '✅', ''),
        ('인프라', '운영 인프라', 'Phase 1.5 — 남은', '개발환경 서버 구축 (Mac mini 자체)', 'Mac mini 24/7 + Docker Postgres + Cloudflare Tunnel', '⬜', '메모리 권장 셋업'),
        ('인프라', '운영 인프라', 'Phase 1.5 — 남은', '통합 테스트 (3콘솔 + mobile 실기기)', '실 서버 + 실 키 + 3콘솔 + mobile walk-through', '⬜', ''),
        ('인프라', '운영 인프라', '남은', '운영 서버 배포', '개발환경 검증 → 운영 전환 (Mac mini 또는 가비아 호스팅)', '⬜', '심의 제출 직전'),

        # ════════ 자동화 (Mac mini 자체구축) ════════
        ('인프라', '자동화', '남은 (Stage 1)', '카톡 챗봇 — 자체 webhook', '카카오 i 오픈빌더 무료 + Mac mini webhook 처리', '⬜', '메모리 — Mac mini 자체'),
        ('인프라', '자동화', '남은 (Stage 1)', '이메일 자동 응답 — 자체 분류', 'Mac mini Python 스크립트 + IMAP 폴링', '⬜', '메모리 services/customer_support_ai.py'),
        ('인프라', '자동화', '남은 (Stage 2)', 'AI 응답 (자체 LLM 또는 OpenAI)', 'Mac mini Llama 양자화 또는 ChatGPT API (소규모)', '⬜', '대규모는 외부 API'),
        ('인프라', '자동화', '남은 (Stage 2)', 'SNS 자동 게시 — 자체 cron', 'Mac mini cron + Meta/Insta/YouTube/TikTok API 직접', '⬜', 'Buffer 등 SaaS 대체'),
        ('인프라', '자동화', '남은 (Stage 2)', '앱 스토어 리뷰 폴링 + AI 응답', 'Mac mini services/app_store_reviews.py + review_response_ai.py', '⬜', '메모리'),
        ('인프라', '자동화', '남은 (Stage 3)', '사용자 행동 시퀀스', '백엔드 services/notification_sequence.py', '⬜', '출시 후'),
        ('슈퍼어드민', '자동화', '남은 (Stage 2)', 'admin 자동화 대시보드', '고객응대/마케팅 현황 — admin-web 메뉴 추가', '⬜', 'Phase D'),

        # ════════ 후속 단계 (Phase 2~5) ════════
        ('인프라', 'Phase 2~4', '남은', 'W7 — 6 페르소나 테스트 시드', '외국인/한국인/소규모/중대형/직원/슈퍼어드민', '⬜', '7월 중~하순'),
        ('인프라', 'Phase 5', '남은', '스토어 메타데이터 + 빌드', 'iOS/Android 빌드 + 스크린샷 + 메타데이터(다국어) + 심의', '⬜', '7월 하순~8월 초'),
    ]

    last_domain = ''
    for it in items:
        ws.append(it)
        r = ws.max_row
        actor_color = ACTOR_COLOR.get(it[0])
        if actor_color:
            ws.cell(row=r, column=1).fill = PatternFill('solid', start_color=actor_color)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=1).font = Font(name=FONT, bold=True)
        if it[1] != last_domain:
            ws.cell(row=r, column=2).font = Font(name=FONT, bold=True)
            ws.cell(row=r, column=2).fill = SUBHEADER_FILL
            last_domain = it[1]
        status_fill = {'✅': DONE_FILL, '◑': PARTIAL_FILL, '🔄': ACTIVE_FILL, '⬜': TODO_FILL}.get(it[5])
        if status_fill:
            ws.cell(row=r, column=6).fill = status_fill
            ws.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center')

    apply_table(ws, header_row, ws.max_row, len(headers))
    for i, w in enumerate([14, 22, 18, 36, 60, 8, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 28

    # ── Phase 1 PR 체크리스트 ──
    ws2 = wb.create_sheet('Phase 1 PR 체크리스트')
    pr_headers = ['PR', '분류', '주체 영향', '도메인', '내용 요약',
                  '상태', '진척율', '담당', '시작일', '목표일', '완료일', '비고']
    ws2.append(pr_headers)
    prs = [
        ('P1', '인프라', '사용자', 'mobile 디자인 기반', '테마 단일화 + 시스템 폰트 + Pw* 4종', '✅', '1인', '2026-05-21', '2026-05-22', today, ''),
        ('P2', '인프라', '사용자·Pathwave', 'mobile i18n', '12언어 + 전 화면 t() + ko 시드 550 + DeepL 스크립트', '✅', '1인', '2026-05-21', '2026-05-22', today, ''),
        ('P3', '인프라', '시설관리자·슈퍼어드민', '웹 디자인 시스템', 'alert/confirm 33곳 → useDialog + 색 토큰', '✅', '1인', '2026-05-21', '2026-05-22', today, ''),
        ('P4', 'Critical', '시설관리자·Pathwave', '인증 우회 제거', 'DEV_AUTO_LOGIN env + 실 로그인폼 + /forgot-password 정식', '✅', '1인', today, today, today, '⚠ app.py 재시작'),
        ('P5', 'Critical', '시설관리자', '매장·회사정보 실연동', 'StoreInfo 실연동 + PwFooter + dead code 정리', '✅', '1인', today, today, today, ''),
        ('P6', 'Critical', '시설관리자', 'OCR 허위 제거', '가짜 OCR 자동입력 → 정직한 수동입력 UI', '✅', '1인', today, today, today, ''),
        ('P7', 'Critical', '시설관리자', '결제·구독 실연동', '카드 localStorage 제거(PCI) + BillingService + 실연동', '✅', '1인', today, today, today, ''),
        ('P8', 'Critical', '전체', '채팅 도메인', 'provider 실연동 + admin BE + mobile SSE 재연결', '◑', '1인', today, today, today, '번역=P8b, ⚠ 재시작'),
        ('P8b', 'Critical', '전체', '채팅 자동 번역', '번역 캐시 + 뷰어별 언어 + translator 연결', '⬜', '1인', '', '키 확보 후', '', 'DeepL 키 후'),
        ('P9', 'Critical', '사용자·시설관리자', '쿠폰·스탬프 실연동', 'mock→실연동 + 사용 silent error', '⬜', '1인', '', '2026-06-중', '', ''),
        ('P10', 'Critical', '전체', '알림 도메인', 'provider 실연동 + mobile 라우팅 + 웹푸시 자체구축', '⬜', '1인', '', '2026-06-중', '', ''),
        ('P11', 'Critical', '시설관리자·슈퍼어드민', '대시보드·직원', 'Dashboard·StaffManagement·CouponStats 실연동', '⬜', '1인', '', '2026-06-중', '', ''),
        ('P12', 'Critical', '사용자', 'mobile 심의 직격', 'consent placeholder + dev 정보 제거 + 버전 동적화', '⬜', '1인', '', '2026-06-중', '', ''),
        ('P13', 'Critical', '전체', '약관 3종', '환불·청소년·쿠키 정책', '⬜', '1인', '', '2026-06-중', '', ''),
        ('P14', 'WiFi 로밍 B', 'Pathwave', 'WiFi 데이터 모델', 'wifi_profiles 확장 + 신규 테이블 + beacons.role', '⬜', '1인', '', '6/말~7/초', '', ''),
        ('P15', 'WiFi 로밍 B', '시설관리자·슈퍼어드민', 'WiFi 등록·연동', 'handshake 묶음 + admin 등록 + provider 실연동', '⬜', '1인', '', '7/초', '', ''),
        ('P16', 'WiFi 로밍 B', '사용자', 'mobile WiFi 클라이언트', '비콘→WiFi 묶음 + BSSID 검증', '⬜', '1인', '', '7/초', '', ''),
        ('P17', 'WiFi 로밍 B', '사용자', '.mobileconfig 다건', '다건 설치', '⬜', '1인', '', '7/중', '', '서명은 인증서 후'),
        ('P18', 'WiFi 로밍 B', '시설관리자·사용자', 'credential_mode managed', '비번 교체 + 자동 전파', '⬜', '1인', '', '7/중', '', 'v1 flag'),
        ('P19', 'WiFi 로밍 B', '슈퍼어드민·시설관리자', 'units/grant 관리', '호실·자리 권한', '⬜', '1인', '', '7/중', '', 'v1 flag'),
        ('P20', '심의 마무리', '슈퍼어드민·사용자', '앱 버전관리', 'admin app-versions UI + mobile 게이트', '⬜', '1인', '', '7/중', '', ''),
        ('P21', '심의 마무리', '인프라', '심의 메타 자산', 'PrivacyInfo + Bundle ID + Photo Picker', '⬜', '1인', '', '7/중', '', ''),
        ('P22', '회원 QR', '사용자·시설관리자', '쿠폰·스탬프 회원 QR', '회원 QR + provider 스캔/코드입력', '⬜', '1인', '', '7/초~중', '', ''),
    ]
    for row in prs:
        ws2.append(row)
    for r in range(2, len(prs) + 2):
        status_cell = ws2.cell(row=r, column=6).coordinate
        ws2.cell(row=r, column=7).value = (
            f'=IF({status_cell}="✅",1,IF({status_cell}="🔄",0.5,'
            f'IF({status_cell}="◑",0.7,IF({status_cell}="🔎",0.9,0))))'
        )
        ws2.cell(row=r, column=7).number_format = PCT_FMT
        status = ws2.cell(row=r, column=6).value
        fill = {'✅': DONE_FILL, '◑': PARTIAL_FILL, '🔄': ACTIVE_FILL, '🔎': ACTIVE_FILL, '⬜': TODO_FILL}.get(status)
        if fill:
            ws2.cell(row=r, column=6).fill = fill
            ws2.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center')
        for col in (9, 10, 11):
            ws2.cell(row=r, column=col).number_format = DATE_FMT
            ws2.cell(row=r, column=col).font = INPUT_FONT
        ws2.cell(row=r, column=8).font = INPUT_FONT
    dv = DataValidation(type='list', formula1='"⬜,🔄,🔎,◑,✅"', allow_blank=True)
    dv.add(f'F2:F{len(prs)+1}')
    ws2.add_data_validation(dv)
    apply_table(ws2, 1, len(prs) + 1, len(pr_headers))
    for i, w in enumerate([6, 12, 22, 22, 46, 8, 9, 9, 12, 18, 12, 28], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 28

    # ── 단계별 일정 ──
    ws3 = wb.create_sheet('단계별 일정')
    ws3.append(['단계', '내용', '예상 기간', '누적', '비고'])
    schedule = [
        ('이전 작업 (~Phase 0)', 'PR #1~#51 + 디자인 통합', '~수개월', '2026-05-21', '완료'),
        ('Phase 0 전수 감사', '갭 14 Critical + High 발굴', '~1일', '2026-05-21', '완료'),
        ('Phase 1 인프라', 'P1~P3', '~4일', '2026-05-22', '완료'),
        ('Phase 1 Critical', 'P4~P13', '~2주', '2026-06-중', 'P4~P8 완료'),
        ('Phase 1 WiFi 로밍 B', 'P14~P19', '~2.5~3주', '7/초~중', 'flag 비공개'),
        ('Phase 1 회원 QR', 'P22 (P9 이후 병행)', '~1주', '7/초~중', ''),
        ('Phase 1 심의 마무리', 'P20~P21', '~3일', '7/중', ''),
        ('Phase 1.5 — Mac mini 개발환경 구축', 'Mac mini 24/7 + Docker Postgres + Cloudflare Tunnel + 환경변수 키 주입', '~3일', '7/중', '메모리 권장 셋업'),
        ('Phase 1.5 — 통합 테스트', '실 서버 + 실 키 + 3콘솔 + mobile 실기기 walk-through', '~4일', '7/중~하순', ''),
        ('Phase 2~4 페르소나 검증', '6 페르소나 시나리오', '~1주', '7/하순', ''),
        ('Phase 5 제출', '빌드 + 스토어 메타데이터', '며칠', '7/하순', ''),
        ('출시', '앱스토어/Play 공개', '—', '7/하순~8/초', ''),
        ('외부 서비스 신청', '법인카드 수취 후 일괄 (던스 → Apple → 도메인 등)', '~1주 (심사 1~2주)', '2026-06-02 주', '병행'),
    ]
    for row in schedule:
        ws3.append(row)
    apply_table(ws3, 1, len(schedule) + 1, 5)
    for i, w in enumerate([30, 50, 16, 14, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[1].height = 28

    # ── 요약 ──
    ws4 = wb.create_sheet('요약', 0)
    ws4['A1'] = 'PathWave 개발 체크리스트 요약'
    ws4['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
    ws4.merge_cells('A1:C1')
    ws4.append([])
    ws4.append(['지표', '값'])
    for c in (1, 2):
        ws4.cell(row=3, column=c).font = HEADER_FONT
        ws4.cell(row=3, column=c).fill = HEADER_FILL
        ws4.cell(row=3, column=c).alignment = Alignment(horizontal='center')
        ws4.cell(row=3, column=c).border = BORDER
    summary = [
        ('총 PR 수 (Phase 1)',   "=COUNTA('Phase 1 PR 체크리스트'!A2:A30)"),
        ('완료 (✅)',            "=COUNTIF('Phase 1 PR 체크리스트'!F2:F30,\"✅\")"),
        ('부분 완료 (◑)',       "=COUNTIF('Phase 1 PR 체크리스트'!F2:F30,\"◑\")"),
        ('진행중 (🔄)',         "=COUNTIF('Phase 1 PR 체크리스트'!F2:F30,\"🔄\")"),
        ('대기 (⬜)',            "=COUNTIF('Phase 1 PR 체크리스트'!F2:F30,\"⬜\")"),
        ('가중 평균 진척율',    "=AVERAGE('Phase 1 PR 체크리스트'!G2:G30)"),
        ('업데이트 일자',        today),
        ('외부 서비스 신청 시작', '2026-06-02 주'),
        ('Phase 1.5 개발환경+테스트', '2026-07 중~하순'),
        ('목표 출시일',          '2026-07-하순~8월 초'),
    ]
    for i, (label, value) in enumerate(summary):
        r = 4 + i
        ws4.cell(row=r, column=1).value = label
        ws4.cell(row=r, column=2).value = value
        for c in (1, 2):
            ws4.cell(row=r, column=c).border = BORDER
        if label == '가중 평균 진척율':
            ws4.cell(row=r, column=2).number_format = PCT_FMT
            ws4.cell(row=r, column=2).fill = ASSUME_FILL
        if isinstance(value, str) and value.startswith('='):
            ws4.cell(row=r, column=2).font = LINK_FONT
    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 40
    ws4.append([])
    ws4.cell(row=ws4.max_row + 1, column=1).value = '주체 컬러 범례'
    ws4.cell(row=ws4.max_row, column=1).font = Font(name=FONT, bold=True, color='1F4E78')
    for actor, color in ACTOR_COLOR.items():
        ws4.append([actor, ''])
        ws4.cell(row=ws4.max_row, column=1).fill = PatternFill('solid', start_color=color)
        ws4.cell(row=ws4.max_row, column=1).font = Font(name=FONT, bold=True)
        desc = {
            '사용자': 'mobile 앱 손님',
            '시설관리자': 'provider-web 점주',
            '슈퍼어드민': 'admin-web 운영자',
            'Pathwave 서비스': 'Pathwave 백엔드 (API/DB/번역/푸시 등 공통)',
            '인프라': '서버·운영도구·자동화·심의 메타',
        }.get(actor, '')
        ws4.cell(row=ws4.max_row, column=2).value = desc

    out = 'docs/exports/pathwave_dev_checklist.xlsx'
    force_workbook_font(wb)
    wb.save(out)
    return out


# ═══════════════════════════════════════════════════════════════════════════
# FILE 2 — 서비스 신청 + 비용
# ═══════════════════════════════════════════════════════════════════════════
def build_services_costs():
    wb = Workbook()
    wb.remove(wb.active)

    # ── 서비스 신청 체크리스트 ──
    ws = wb.create_sheet('서비스 신청 체크리스트')
    headers = ['우선순위', '카테고리', '서비스', '제공사', '필수/선택',
               '가격 상세', '신청 링크', '가입 계정',
               '신청 상태', '신청 예정', '활성 예정', '환경변수 / 비고']
    ws.append(headers)

    services = [
        # 우선순위 1 — 6/2주 즉시 (법인카드 수취 후)
        (1, '법인 식별',  'D-U-N-S 번호 (Apple 법인 가입 필수)', '한국기업데이터(cretop) / D&B',  '필수', '무료 (5~10영업일)',                                'https://www.cretop.com/',                              'admin@triggersoft', '대기', '2026-06-02', '2026-06-12', 'Apple Developer 법인 계정 선행조건 — 무료 신청'),
        (1, '결제',       '토스페이먼츠',           'Toss Payments',                '필수', '수수료 2.9% + 33원/건 (가입 무료)',                'https://merchant.tosspayments.com/',                  'admin@pathwave',    '대기', '2026-06-02', '2026-06-16', 'TOSS_SECRET_KEY · 심사 1~2주'),
        (1, '스토어',     'Apple Developer Program (법인)', 'Apple',                '필수', '$99/년 (₩148,500 @1500) · D-U-N-S 필요',         'https://developer.apple.com/programs/enroll/',         'admin@pathwave',    '대기', '2026-06-12', '2026-06-19', 'D-U-N-S 발급 후 가입 가능'),
        (1, '스토어',     'Google Play Console',     'Google',                       '필수', '$25 일회성 (₩37,500 @1500)',                       'https://play.google.com/console/u/0/signup',           'admin@pathwave',    '대기', '2026-06-02', '2026-06-03', ''),
        (1, '도메인',     'pathwave.co.kr',          '가비아',                        '필수', '연 ~3만원 (월 ~₩2,500)',                          'https://www.gabia.com/',                               'admin@pathwave',    '대기', '2026-06-02', '2026-06-02', '서비스 도메인'),
        (1, '도메인',     'triggersoft.kr',          '가비아',                        '필수', '연 ~3만원 (월 ~₩2,500)',                          'https://www.gabia.com/',                               'admin@triggersoft', '대기', '2026-06-02', '2026-06-02', '법인 운영'),
        (1, '호스팅(권장)','Mac mini 자체 (Cloudflare Tunnel)','자체 보유',           '필수', '월 ₩0 (전기료 ~₩3,000)',                          'https://dash.cloudflare.com/',                         'admin@pathwave',    '대기', '2026-06-09', '2026-07-중', '메모리 권장 — Mac mini 1대로 백엔드+자동화'),
        (1, '호스팅(대안)','가비아 SimpleHosting Light', '가비아',                    '선택', '월 ~₩4,500 (Python 지원)',                        'https://www.gabia.com/service/hosting',                'admin@pathwave',    '대기', '대안',       '',           'Mac mini 부재 시 대안'),
        (1, 'DB',         'Supabase 또는 Neon (Postgres 무료)','Supabase / Neon',    '필수', '무료 tier (500MB ~ 1GB)',                          'https://supabase.com/',                                'admin@pathwave',    '대기', '2026-06-09', '2026-06-12', '메모리 권장 무료 tier'),
        (1, '인증',       'Firebase Auth',           'Google',                        '필수', '무료 (10K MAU)',                                   'https://console.firebase.google.com/',                 'dev@pathwave',      '대기', '2026-06-02', '2026-06-03', '인증 전용'),
        (1, '푸시(iOS)',  'APNs (자체구축)',         'Apple',                         '필수', '무료 (Apple Developer 포함)',                      'https://developer.apple.com/account/resources/',       'admin@pathwave',    '대기', '2026-06-19', '2026-06-25', 'APNs HTTP/2 + JWT (PR #50 완성)'),
        (1, '푸시(Android)','FCM (무료 무제한)',     'Google',                        '필수', '무료 무제한',                                      'https://console.firebase.google.com/',                 'dev@pathwave',      '대기', '2026-06-02', '2026-06-03', '⚠ Android는 FCM 외 사실상 대안 없음 — Google Play Services 의존'),
        (1, '이메일도메인','Google Workspace (Pathwave)','Google',                    '필수', '$7/계정/월 (₩10,500) + alias 4개 무료',           'https://workspace.google.com/',                        'admin@pathwave',    '대기', '2026-06-02', '2026-06-02', 'admin/dev/support/noreply/info 5메일'),
        (1, '이메일도메인','Google Workspace (Triggersoft)','Google',                 '필수', '$7/계정/월 (₩10,500)',                            'https://workspace.google.com/',                        'admin@triggersoft', '대기', '2026-06-02', '2026-06-02', '법인 운영'),
        (1, '상표등록',   '상표등록 — 패스웨이브',     '특허청 (셀프 또는 변리사)',     '필수', '셀프 ~₩267,000 / 변리사 50~80만',                'https://www.patent.go.kr/',                            'admin@triggersoft', '대기', '2026-06-09', '심사 6~12개월', '출원 56,000 + 등록 211,000 (10년)'),
        (1, '상표등록',   '상표등록 — 트리거소프트',   '특허청 (셀프 또는 변리사)',     '필수', '셀프 ~₩267,000 / 변리사 50~80만',                'https://www.patent.go.kr/',                            'admin@triggersoft', '대기', '2026-06-09', '심사 6~12개월', ''),

        # 우선순위 2 — 6/2주 신청 (심사 1~2주)
        (2, '소셜로그인', '카카오 로그인',            'Kakao',                          '필수', '무료',                                            'https://developers.kakao.com/',                        'info@pathwave',     '대기', '2026-06-02', '2026-06-16', '검수 1~2주'),
        (2, '소셜로그인', '네이버 로그인',            'Naver',                          '필수', '무료',                                            'https://developers.naver.com/',                        'info@pathwave',     '대기', '2026-06-02', '2026-06-09', ''),
        (2, '이메일 발송','SendGrid (무료 → Essentials)', 'Twilio',                  '필수', '무료 100통/일 · Essentials 50K/월 $19.95 (~₩30K)','https://signup.sendgrid.com/',                         'dev@pathwave',      '대기', '2026-06-02', '2026-06-03', '매장 100+ 시 유료 tier 필요'),
        (2, '모니터링',   'Sentry (무료 또는 자체호스팅)', 'Sentry',                    '필수', '무료 5K events/월 · self-hosted 무료',            'https://sentry.io/signup/',                            'dev@pathwave',      '대기', '2026-06-02', '2026-06-03', '대규모 시 자체 호스팅 Mac mini'),
        (2, '지도',       'Google Maps API',         'Google',                        '필수', '월 $200 무료 크레딧',                              'https://console.cloud.google.com/',                    'dev@pathwave',      '대기', '2026-06-02', '2026-06-03', '카드 등록 필수'),

        # 우선순위 3 — 옵션·후속
        (3, 'SMS 인증(옵션)', '솔라피 SMS', 'Solapi',                                  '선택', '건당 8~15원 (LMS 28~35원) · 사용량 기반',          'https://solapi.com/',                                  'dev@pathwave',      '대기', '대안',       '',           '이메일 인증으로 대체 가능 — 매장 100+ 시 도입'),
        (3, '알림톡(옵션)', '솔라피 알림톡', 'Solapi',                                 '선택', '건당 7~9원 · 카카오톡 채널 필요',                  'https://solapi.com/',                                  'dev@pathwave',      '대기', '대안',       '',           '⚠ 자체 앱 푸시 + 이메일로 대체 가능 — 외부 의존 줄이려면 미도입'),
        (3, '본인인증',   'NICE / KCB / PASS',        '본인인증 사업자',                '선택', '보증금 ~10만원 + 건당 30~80원',                    'https://www.niceinfo.co.kr/',                          'admin@pathwave',    '대기', '6/말~7/중', '',           '미성년자 확인 (대체 불가)'),
        (3, '번역',       'DeepL Pro',               'DeepL',                         '선택', '무료 500K chars/월 · Starter ~€5/월',             'https://www.deepl.com/pro',                            'dev@pathwave',      '대기', '2026-07-초', '2026-07-중', 'P8b 채팅 자동번역 USP'),

        # 우선순위 3 — 자동화 (Mac mini 자체구축 기본, 외부는 옵션)
        (3, '자동화(자체)','챗봇·이메일·SNS 게시 — Mac mini 자체구축', 'Mac mini',     '필수', '월 ₩0 (전기 + 개발 시간)',                         '—',                                                    'dev@pathwave',      '대기', '출시 후',    '',           '메모리 권장 — Python 스크립트 + cron + 카카오 i 오픈빌더 webhook'),
        (3, '자동화(옵션)','ChatGPT API (소규모 AI 응답)', 'OpenAI',                  '선택', '$0.15~3/M tokens (사용량)',                       'https://platform.openai.com/signup',                   'dev@pathwave',      '대기', '대안',       '',           'Mac mini Llama 양자화 자체 LLM 대체 가능'),
        (3, '자동화(옵션)','Twilio Voice (AI 음성)',   'Twilio',                       '선택', '분당 100~500원',                                   'https://www.twilio.com/',                              'admin@pathwave',    '대기', '+3개월',     '',           'Stage 2 — 070 회선 후 도입'),
    ]
    for s in services:
        ws.append(s)
    for r in range(2, len(services) + 2):
        link = ws.cell(row=r, column=7).value
        if link and isinstance(link, str) and link.startswith('http'):
            ws.cell(row=r, column=7).hyperlink = link
            ws.cell(row=r, column=7).font = Font(name=FONT, color='0563C1', underline='single')
        for col in (10, 11):
            ws.cell(row=r, column=col).font = INPUT_FONT
        status = ws.cell(row=r, column=9).value
        fill = {'활성': DONE_FILL, '검수중': PARTIAL_FILL, '신청중': ACTIVE_FILL,
                '대기': TODO_FILL, '대안': PARTIAL_FILL, '보류': PARTIAL_FILL}.get(status)
        if fill:
            ws.cell(row=r, column=9).fill = fill
            ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='center')
        pri = ws.cell(row=r, column=1).value
        pri_color = {1: 'FFC7CE', 2: 'FFEB9C', 3: 'F2F2F2'}.get(pri)
        if pri_color:
            ws.cell(row=r, column=1).fill = PatternFill('solid', start_color=pri_color)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=1).font = Font(name=FONT, bold=True)
    dv1 = DataValidation(type='list', formula1='"대기,신청중,검수중,활성,대안,보류"', allow_blank=True)
    dv1.add(f'I2:I{len(services)+1}')
    ws.add_data_validation(dv1)
    apply_table(ws, 1, ws.max_row, len(headers))
    for i, w in enumerate([7, 14, 38, 28, 9, 36, 38, 18, 10, 13, 13, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36

    # ── 초기 세팅비 ──
    ws_init = wb.create_sheet('초기 세팅비 (1회성)')
    ws_init['A1'] = '초기 세팅비 — 출시 전 1회성 / 연 1회 (월별 시트에 분산 안 함)'
    ws_init['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
    ws_init.merge_cells('A1:D1')
    ws_init.append([])
    ws_init.append(['항목', '금액 (₩)', '주기', '비고'])
    init_items = [
        ('D-U-N-S 번호 신청 (한국기업데이터)',          0, '일회성', '무료 — Apple 법인 가입 선행, 5~10영업일'),
        ('Apple Developer Program — 첫 회비',     148500, '연 1회', '$99 × 1500'),
        ('Google Play Console — 등록비',           37500, '일회성', '$25 일회성'),
        ('도메인 — pathwave.co.kr',                30000, '연 1회', '서비스 도메인 (가비아)'),
        ('도메인 — triggersoft.kr',                30000, '연 1회', '법인 도메인 (가비아)'),
        ('상표등록 — 패스웨이브 (셀프 출원)',      267000, '일회성', '출원 56,000 + 등록 211,000 (10년 갱신)'),
        ('상표등록 — 트리거소프트 (셀프 출원)',     267000, '일회성', '법인명 상표'),
        ('Mac mini M1 16GB 중고 (운영 베이스)',    800000, '일회성', '자체 호스팅 + 자동화 — 메모리 권장'),
        ('Mac mini 주변기기 (외장 SSD)',            80000, '일회성', 'Mac mini 옵션'),
        ('NICE 본인인증 — 보증금 (선택)',          100000, '일회성', '미성년자 확인 도입 시'),
        ('카카오 / 네이버 로그인 검수',                 0, '일회성', '무료, 1~2주'),
        ('솔라피 알림톡 템플릿 등록 (도입 시)',         0, '일회성', '무료, 1~2주'),
        ('토스페이먼츠 가맹점 가입',                    0, '일회성', '무료, 심사 1~2주'),
        ('사업자등록 / 통신판매업 / 위치기반신고',      0, '일회성', '무료'),
        ('Cloudflare / Firebase / Sentry 가입',         0, '일회성', '무료'),
    ]
    for it in init_items:
        ws_init.append(it)
        r = ws_init.max_row
        ws_init.cell(row=r, column=2).font = INPUT_FONT
        ws_init.cell(row=r, column=2).number_format = KRW_FMT
    last_init = ws_init.max_row
    ws_init.append(['소계 — 필수만 (Mac mini·NICE 제외)',
                    f'=SUM(B4:B9)+B14+SUM(B15:B{last_init})', '', 'Apple/Play/도메인/상표만'])
    ws_init.append(['소계 — Mac mini 포함 (메모리 권장)',
                    f'=SUM(B4:B9)+SUM(B11:B12)+B14+SUM(B15:B{last_init})', '', '+ Mac mini 880,000'])
    ws_init.append(['소계 — 전체 (옵션 모두 포함)',
                    f'=SUM(B4:B{last_init})', '', '+ NICE 100,000'])
    for r in (ws_init.max_row - 2, ws_init.max_row - 1, ws_init.max_row):
        ws_init.cell(row=r, column=2).number_format = KRW_FMT
        ws_init.cell(row=r, column=2).font = Font(name=FONT, bold=True)
        ws_init.cell(row=r, column=2).fill = ASSUME_FILL
        ws_init.cell(row=r, column=1).font = Font(name=FONT, bold=True)
        ws_init.cell(row=r, column=1).fill = ASSUME_FILL
    apply_table(ws_init, 3, ws_init.max_row, 4)
    for i, w in enumerate([48, 16, 12, 50], 1):
        ws_init.column_dimensions[get_column_letter(i)].width = w

    # ── 월별 운영비 시나리오 (매장 수 기반) ──
    ws2 = wb.create_sheet('월별 운영비 시나리오')
    ws2['A1'] = '월별 운영비 시나리오 — 매장 수 기반 (자체구축 우선, 외부는 옵션)'
    ws2['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
    ws2.merge_cells('A1:F1')
    ws2.append([])
    # 가정 입력 행 (2줄)
    ws2.append(['매장당 평균 사용자 (가정)', 300, '', '매장당 매출 (가정, ₩/월)', 10000, ''])
    ws2.append(['매장당 비콘 개수 (가정)',   2,   '', '비콘 단가 (₩/개, 배송 포함)', 10000, ''])
    for r in (3, 4):
        for c in (2, 5):
            ws2.cell(row=r, column=c).font = INPUT_FONT
            ws2.cell(row=r, column=c).fill = ASSUME_FILL
        ws2.cell(row=r, column=5).number_format = KRW_FMT
        for c in (1, 4):
            ws2.cell(row=r, column=c).font = Font(name=FONT, bold=True)
    ws2.append([])

    ws2.append(['항목', '매장 10 (초기)', '매장 10,000 (Y1 목표)', '매장 30,000 (Y2~3)', '매장 70,000 (Y3+)', '비고'])
    header_row2 = ws2.max_row
    # 정보 행 (매장수·MAU·매출·비콘 CAPEX)
    ws2.append(['매장 수',                  10,         10000,        30000,        70000,        ''])
    ws2.append(['MAU (매장×평균)',           '=B7*$B$3', '=C7*$B$3',   '=D7*$B$3',   '=E7*$B$3',   ''])
    ws2.append(['월 매출 (매장×매장당매출)',  '=B7*$E$3', '=C7*$E$3',   '=D7*$E$3',   '=E7*$E$3',   ''])
    ws2.append(['비콘 누적 CAPEX (매장×개수×단가)',
                '=B7*$B$4*$E$4', '=C7*$B$4*$E$4', '=D7*$B$4*$E$4', '=E7*$B$4*$E$4',
                '매장 가입 시 1회 — 매장당 매출 ₩10K 기준 2개월 회수'])
    for r in (7, 8, 9, 10):
        for c in (2, 3, 4, 5):
            if r == 7:
                ws2.cell(row=r, column=c).font = INPUT_FONT
                ws2.cell(row=r, column=c).number_format = '#,##0'
            elif r == 8:
                ws2.cell(row=r, column=c).font = LINK_FONT
                ws2.cell(row=r, column=c).number_format = '#,##0'
            else:
                ws2.cell(row=r, column=c).font = LINK_FONT
                ws2.cell(row=r, column=c).number_format = KRW_FMT
        ws2.cell(row=r, column=1).font = Font(name=FONT, bold=True)
        if r == 10:
            for c in (2, 3, 4, 5):
                ws2.cell(row=r, column=c).fill = PatternFill('solid', start_color='FFE4E1')
    # 비용 항목
    cost_start = ws2.max_row + 1
    scen = [
        ('서버 (Mac mini 자체, 전기)',           3000,   3000,   3000,   3000,   '메모리 권장 — Cloudflare Tunnel'),
        ('서버 확장 (대안: 가비아/VPS)',          0,      0,      30000,  100000, '매장 500+ 시 부하 분산 필요 시'),
        ('도메인 (2개 × 월할)',                   5000,   5000,   5000,   5000,   'pathwave + triggersoft'),
        ('Workspace (2계정 × $7 = ₩21K)',         21000,  21000,  21000,  21000,  ''),
        ('이메일 (SendGrid)',                     0,      30000,  30000,  135000, '무료→Essentials $20→Pro $90 (매장 1K=10K통/월 이상)'),
        ('SMS 인증 (옵션 — 이메일 인증 대체 가능)', 0,    5000,   30000,  60000,  '솔라피 — 또는 이메일 인증으로 대체'),
        ('알림톡 (옵션 — 자체 푸시로 대체 가능)',   0,    0,      0,      0,      '⚠ 카카오톡 채널 운영 시만 — 자체 푸시/이메일이면 0'),
        ('Firebase Auth (10K MAU 무료)',          0,      0,      0,      50000,  '매장 1K(MAU 50K)에서 일부 유료'),
        ('Google Maps (월 $200 무료)',            0,      0,      0,      30000,  '대규모 트래픽 시 초과분'),
        ('모니터링 (Sentry 무료 / 자체)',         0,      0,      0,      50000,  '자체 호스팅 시 0'),
        ('번역 API (P8b)',                        0,      0,      30000,  100000, 'DeepL — 채팅 자동 번역'),
        ('자동화 (Mac mini 자체)',                0,      0,      0,      0,      '메모리 — Python 스크립트'),
    ]
    for row in scen:
        ws2.append(row)
        r = ws2.max_row
        for c in (2, 3, 4, 5):
            ws2.cell(row=r, column=c).number_format = KRW_FMT
            ws2.cell(row=r, column=c).font = INPUT_FONT
    cost_end = ws2.max_row
    ws2.append(['월 운영비 합계',
                f'=SUM(B{cost_start}:B{cost_end})',
                f'=SUM(C{cost_start}:C{cost_end})',
                f'=SUM(D{cost_start}:D{cost_end})',
                f'=SUM(E{cost_start}:E{cost_end})',
                ''])
    total_row = ws2.max_row
    for c in (2, 3, 4, 5):
        ws2.cell(row=total_row, column=c).number_format = KRW_FMT
        ws2.cell(row=total_row, column=c).font = Font(name=FONT, bold=True)
        ws2.cell(row=total_row, column=c).fill = ASSUME_FILL
    ws2.cell(row=total_row, column=1).font = Font(name=FONT, bold=True)
    ws2.cell(row=total_row, column=1).fill = ASSUME_FILL
    # 손익
    ws2.append(['월 손익 (매출 − 비용)',
                f'=B9-B{total_row}',
                f'=C9-C{total_row}',
                f'=D9-D{total_row}',
                f'=E9-E{total_row}',
                ''])
    pl_row = ws2.max_row
    for c in (2, 3, 4, 5):
        ws2.cell(row=pl_row, column=c).number_format = KRW_FMT
        ws2.cell(row=pl_row, column=c).font = Font(name=FONT, bold=True, color='C00000')
        ws2.cell(row=pl_row, column=c).fill = ASSUME_FILL
    ws2.cell(row=pl_row, column=1).font = Font(name=FONT, bold=True, color='C00000')
    ws2.cell(row=pl_row, column=1).fill = ASSUME_FILL

    apply_table(ws2, header_row2, pl_row, 6)
    for i, w in enumerate([38, 18, 18, 18, 18, 38], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── 손익분기점 (BEP) ──
    ws_bep = wb.create_sheet('손익분기점 (BEP)')
    ws_bep['A1'] = '손익분기점 — 매장 수가 매출 단위'
    ws_bep['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
    ws_bep.merge_cells('A1:E1')
    ws_bep.append([])
    ws_bep.append(['시나리오', '월 운영비', '매장 수 (실제)', '월 매출', '월 손익'])
    ws_bep.append(['매장 10 (초기)',         f"='월별 운영비 시나리오'!B{total_row}", 10,    "=C4*'월별 운영비 시나리오'!$E$3", '=D4-B4'])
    ws_bep.append(['매장 10,000 (Y1 목표)',  f"='월별 운영비 시나리오'!C{total_row}", 10000, "=C5*'월별 운영비 시나리오'!$E$3", '=D5-B5'])
    ws_bep.append(['매장 30,000 (Y2~3)',     f"='월별 운영비 시나리오'!D{total_row}", 30000, "=C6*'월별 운영비 시나리오'!$E$3", '=D6-B6'])
    ws_bep.append(['매장 70,000 (Y3+)',      f"='월별 운영비 시나리오'!E{total_row}", 70000, "=C7*'월별 운영비 시나리오'!$E$3", '=D7-B7'])
    for r in (3,):
        for c in (1, 2, 3, 4, 5):
            ws_bep.cell(row=r, column=c).font = HEADER_FONT
            ws_bep.cell(row=r, column=c).fill = HEADER_FILL
            ws_bep.cell(row=r, column=c).alignment = Alignment(horizontal='center')
            ws_bep.cell(row=r, column=c).border = BORDER
    for r in (4, 5, 6, 7):
        for c in (1, 2, 3, 4, 5):
            ws_bep.cell(row=r, column=c).border = BORDER
        ws_bep.cell(row=r, column=2).font = LINK_FONT
        ws_bep.cell(row=r, column=2).number_format = KRW_FMT
        ws_bep.cell(row=r, column=3).font = INPUT_FONT
        ws_bep.cell(row=r, column=4).number_format = KRW_FMT
        ws_bep.cell(row=r, column=5).number_format = KRW_FMT
        ws_bep.cell(row=r, column=5).font = Font(name=FONT, bold=True)
        ws_bep.cell(row=r, column=5).fill = ASSUME_FILL
    ws_bep.append([])
    ws_bep.append(['BEP 매장 수 (운영비 ÷ 매장당 매출)', '', '', '', ''])
    ws_bep.cell(row=ws_bep.max_row, column=1).font = Font(name=FONT, bold=True, color='1F4E78')
    ws_bep.append(['매장 10 시나리오 — BEP',         "=B4/'월별 운영비 시나리오'!$E$3", '', '', '운영비 ÷ 1만원'])
    ws_bep.append(['매장 10,000 시나리오 — BEP',     "=B5/'월별 운영비 시나리오'!$E$3", '', '', ''])
    ws_bep.append(['매장 30,000 시나리오 — BEP',     "=B6/'월별 운영비 시나리오'!$E$3", '', '', ''])
    ws_bep.append(['매장 70,000 시나리오 — BEP',     "=B7/'월별 운영비 시나리오'!$E$3", '', '', ''])
    for r in range(ws_bep.max_row - 3, ws_bep.max_row + 1):
        for c in (1, 2, 3, 4, 5):
            ws_bep.cell(row=r, column=c).border = BORDER
        ws_bep.cell(row=r, column=2).number_format = '#,##0'
        ws_bep.cell(row=r, column=2).font = Font(name=FONT, bold=True)
        ws_bep.cell(row=r, column=2).fill = ASSUME_FILL
    ws_bep.column_dimensions['A'].width = 36
    ws_bep.column_dimensions['B'].width = 18
    ws_bep.column_dimensions['C'].width = 14
    ws_bep.column_dimensions['D'].width = 18
    ws_bep.column_dimensions['E'].width = 18

    ws_bep.append([])
    ws_bep.cell(row=ws_bep.max_row + 1, column=1).value = '⚠ 수익 구조 분석'
    ws_bep.cell(row=ws_bep.max_row, column=1).font = Font(name=FONT, bold=True, color='C00000')
    for n in [
        '· 매장당 월 ₩10,000 매출 — Y1 목표 100매장 = ₩1M/월 매출, 운영비 ~₩60K → 순익 ₩940K/월',
        '· Mac mini 자체구축 + 자체 푸시·이메일로 운영비 대폭 절감',
        '· 알림톡(카카오) 제거 시 외부 의존 0 — 단 솔라피 SMS 인증은 매장 100+ 시 도입 권장(또는 이메일 인증)',
        '· Firebase Auth는 10K MAU 무료 — 매장 1K (MAU 50K) 도달 시 일부 유료 전환',
        '· 자동화는 Mac mini 자체 Python 스크립트 (메모리 권장) — 외부 SaaS 비용 0',
        '· Stage 2~3 도구(Twilio·번역·자동화 SaaS)는 매출 5~10% 재투자 원칙',
        '· 초기 세팅비 회수 — 매장 100 시나리오 순익 ₩940K/월 → 셋업비(필수) ~₩780K 1개월 회수',
    ]:
        ws_bep.append([n])
        ws_bep.cell(row=ws_bep.max_row, column=1).font = Font(name=FONT, color='505050')

    # ── 월별 시트 빌더 (반복만) ──
    def year_sheet(name, year, recurring_data, assumptions_note):
        ws = wb.create_sheet(name)
        ws['A1'] = f'{year}년 월별 운영비 (반복 비용만 — 1회성은 초기 세팅비 시트)'
        ws['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
        ws.merge_cells('A1:N1')
        ws['A2'] = f'※ {assumptions_note}'
        ws['A2'].font = Font(name=FONT, italic=True, color='808080')
        ws.merge_cells('A2:N2')
        ws.append([])
        head = ['항목'] + [f'{m}월' for m in range(1, 13)] + ['연 합계']
        ws.append(head)
        header_row = ws.max_row

        rec_labels = [
            '서버 (Mac mini 자체)', '도메인 (월할)', 'Workspace (2계정)',
            '이메일 (SendGrid)', 'SMS 인증 (옵션)',
            'Firebase Auth (초과)', 'Google Maps (초과)', 'Sentry',
            '번역 API (P8b)', '서버 확장 (대안)',
        ]
        for label, monthly in zip(rec_labels, recurring_data):
            ws.append([label] + list(monthly))
            r = ws.max_row
            ws.cell(row=r, column=14).value = f'=SUM(B{r}:M{r})'
            for c in range(2, 14):
                ws.cell(row=r, column=c).number_format = KRW_FMT
                ws.cell(row=r, column=c).font = INPUT_FONT
            ws.cell(row=r, column=14).number_format = KRW_FMT
            ws.cell(row=r, column=14).font = Font(name=FONT, bold=True)
        last_item = ws.max_row

        ws.append(['월 운영비 합계'] +
                  [f'=SUM({get_column_letter(c)}{header_row+1}:{get_column_letter(c)}{last_item})'
                   for c in range(2, 14)] +
                  [f'=SUM(B{last_item+1}:M{last_item+1})'])
        total_r = ws.max_row
        for c in range(2, 15):
            ws.cell(row=total_r, column=c).number_format = KRW_FMT
            ws.cell(row=total_r, column=c).font = Font(name=FONT, bold=True)
            ws.cell(row=total_r, column=c).fill = ASSUME_FILL
        ws.cell(row=total_r, column=1).font = Font(name=FONT, bold=True)
        ws.cell(row=total_r, column=1).fill = ASSUME_FILL

        ws.append(['누계'] +
                  [f'={get_column_letter(c)}{total_r}' if c == 2
                   else f'={get_column_letter(c-1)}{total_r+1}+{get_column_letter(c)}{total_r}'
                   for c in range(2, 14)] +
                  [f'=N{total_r}'])
        cum_r = ws.max_row
        for c in range(2, 15):
            ws.cell(row=cum_r, column=c).number_format = KRW_FMT
            ws.cell(row=cum_r, column=c).font = LINK_FONT
        ws.cell(row=cum_r, column=1).font = Font(name=FONT, italic=True)

        apply_table(ws, header_row, cum_r, 14)
        ws.column_dimensions['A'].width = 28
        for c in range(2, 15):
            ws.column_dimensions[get_column_letter(c)].width = 11
        return total_r

    # 2026 — 1~5월 코드, 6월부터 활성화, 8월 출시 (매장 0→10)
    R2026 = [
        # 서버 (Mac mini 전기)
        [0,0,0,0,0,0, 3000, 3000, 3000, 3000, 3000, 3000],
        # 도메인 (월할)
        [0,0,0,0,0,5000, 5000, 5000, 5000, 5000, 5000, 5000],
        # Workspace
        [0,0,0,0,0,21000, 21000, 21000, 21000, 21000, 21000, 21000],
        # 이메일 (출시 후 무료 tier 내)
        [0]*12,
        # SMS 인증 (옵션 — 출시 후 소량)
        [0,0,0,0,0,0, 0, 1000, 2000, 3000, 5000, 5000],
        # Firebase Auth (무료 tier)
        [0]*12,
        # Maps (무료 tier)
        [0]*12,
        # Sentry (무료)
        [0]*12,
        # 번역 (P8b 미가동)
        [0]*12,
        # 서버 확장 (불필요)
        [0]*12,
    ]
    total_r_2026 = year_sheet('2026년 월별', 2026, R2026,
        '6/2주 외부 신청 시작. Mac mini 7월 셋업 (자체 호스팅 + Cloudflare Tunnel). 8월 출시 후 매장 0→10. '
        '알림톡/카카오톡 미사용 — 자체 푸시·이메일로 대체')

    # 2027 — 분기별 매장 증가 (Q1 매장 30, Q2 100, Q3 300, Q4 500)
    R2027 = [
        # 서버 (Mac mini 전기 — 일관)
        [3000]*12,
        # 도메인 (월할)
        [5000]*12,
        # Workspace
        [21000]*12,
        # 이메일 — Q2 Essentials 도입 ($20 = ₩30K)
        [0,0,0, 30000,30000,30000, 30000,30000,30000, 30000,30000,30000],
        # SMS — 매장 증가 비례
        [5000,8000,12000, 15000,18000,20000, 25000,30000,35000, 40000,45000,50000],
        # Firebase Auth (Q3+ 일부 유료)
        [0]*8 + [0, 10000, 20000, 30000],
        # Maps
        [0]*10 + [10000, 20000],
        # Sentry (Q4 일부 유료)
        [0]*10 + [0, 30000],
        # 번역 (Q2부터 P8b 가동)
        [0,0,0, 10000,15000,20000, 25000,30000,40000, 50000,60000,80000],
        # 서버 확장 (Q3부터 부하 대비)
        [0]*6 + [0, 30000, 30000, 50000, 50000, 50000],
    ]
    total_r_2027 = year_sheet('2027년 월별', 2027, R2027,
        'Q1 매장 ~30, Q2 ~100, Q3 ~300, Q4 ~500. Mac mini 자체 호스팅 + 자동화. 이메일 Q2부터 SendGrid Essentials. '
        '알림톡 미사용 — 자체 푸시 + 이메일')

    # ── 연간 합계 ──
    ws_sum = wb.create_sheet('연간 합계', 0)
    ws_sum['A1'] = 'PathWave 운영비 — 연간 합계 (반복 + 초기 1회성 분리)'
    ws_sum['A1'].font = Font(name=FONT, bold=True, size=14, color='1F4E78')
    ws_sum.merge_cells('A1:D1')
    ws_sum.append([])
    ws_sum.append(['항목', '금액 (₩)', '월 평균 (₩)', '비고'])
    ws_sum.append(['2026 반복 운영비 (8~12월 5개월)', f"='2026년 월별'!N{total_r_2026}", '=B4/5', '월별 반복만 (매장 0→10)'])
    ws_sum.append(['2027 반복 운영비 (12개월)',       f"='2027년 월별'!N{total_r_2027}", '=B5/12', '월별 반복만 (매장 30→500)'])
    ws_sum.append(['초기 세팅비 — 필수만',             "='초기 세팅비 (1회성)'!B19", '—', 'Apple/Play/도메인/상표만 (Mac mini 제외)'])
    ws_sum.append(['초기 세팅비 — Mac mini 포함 (권장)', "='초기 세팅비 (1회성)'!B20", '—', '+ Mac mini ₩880,000'])
    ws_sum.append(['초기 세팅비 — 전체 (NICE 포함)',    "='초기 세팅비 (1회성)'!B21", '—', '+ NICE ₩100,000'])
    ws_sum.append(['2년 운영비 합계 (반복)',           '=B4+B5', '—', ''])
    ws_sum.append(['2년 총합 (반복 + Mac mini 포함)',  '=B4+B5+B7', '—', '권장 셋업 기준'])
    for r in (3, 4, 5, 6, 7, 8, 9, 10):
        for c in (1, 2, 3, 4):
            ws_sum.cell(row=r, column=c).border = BORDER
    for c in (1, 2, 3, 4):
        ws_sum.cell(row=3, column=c).font = HEADER_FONT
        ws_sum.cell(row=3, column=c).fill = HEADER_FILL
        ws_sum.cell(row=3, column=c).alignment = Alignment(horizontal='center')
    for r in (4, 5, 6, 7, 8, 9, 10):
        ws_sum.cell(row=r, column=2).number_format = KRW_FMT
        ws_sum.cell(row=r, column=2).font = LINK_FONT
        ws_sum.cell(row=r, column=3).number_format = KRW_FMT
    ws_sum.cell(row=10, column=1).font = Font(name=FONT, bold=True)
    ws_sum.cell(row=10, column=2).fill = ASSUME_FILL
    ws_sum.cell(row=10, column=2).font = Font(name=FONT, bold=True)
    ws_sum.column_dimensions['A'].width = 32
    ws_sum.column_dimensions['B'].width = 18
    ws_sum.column_dimensions['C'].width = 18
    ws_sum.column_dimensions['D'].width = 50

    ws_sum.append([])
    ws_sum.cell(row=ws_sum.max_row + 1, column=1).value = '주요 가정 / 변경점 (2026-05-23 v5)'
    ws_sum.cell(row=ws_sum.max_row, column=1).font = Font(name=FONT, bold=True, color='1F4E78')
    for n in [
        '① 환율 USD 1 = ₩1,500',
        '② 서버 = Mac mini 자체 (메모리 권장 인프라 로드맵) — Cloudflare Tunnel로 외부 노출. 비용 거의 0',
        '③ 알림톡 미사용 — 자체 앱 푸시(APNs/FCM) + 이메일로 대체. 외부 카카오톡 의존 0',
        '④ SMS 인증 옵션 — 매장 100+ 시 도입 권장 (또는 이메일 인증으로 대체)',
        '⑤ 자동화 = Mac mini 자체구축 (Python 스크립트 + cron) — Channeltalk/Make/Buffer 등 SaaS 대체',
        '⑥ 던스(D-U-N-S) 무료 신청 — Apple Developer 법인 가입 선행조건 (5~10영업일)',
        '⑦ 도메인·Workspace 2계정 분리 (pathwave + triggersoft)',
        '⑧ 시나리오 = 매장 수 기반 (매장당 매출 ₩10,000 직결). MAU는 매장×평균사용자 자동 계산',
        '⑨ 1회성 비용은 월별 시트 제외 → 초기 세팅비 시트만 (Mac mini 옵션 명시)',
        '⑩ Phase 1.5 단계 — Mac mini 개발환경 구축 + 통합 테스트 (운영 반영 전 필수)',
        '⑪ ⚠ Android 푸시는 FCM 외 사실상 대안 없음 (Google Play Services 의존) — FCM 무료라 실 비용 0',
    ]:
        ws_sum.append([n])
        ws_sum.cell(row=ws_sum.max_row, column=1).font = Font(name=FONT, color='505050')

    out = 'docs/exports/pathwave_services_and_costs.xlsx'
    force_workbook_font(wb)
    wb.save(out)
    return out


if __name__ == '__main__':
    f1 = build_dev_checklist()
    f2 = build_services_costs()
    print('OK:', f1)
    print('OK:', f2)
