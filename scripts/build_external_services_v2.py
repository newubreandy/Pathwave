"""PathWave 외부 서비스 체크리스트 v2 (2026-06-09).

이번 세션 결정 모두 반영:
- 결제 PG 폴백 (제로페이 + 토스)
- OCR = 디바이스 (Apple Vision + ML Kit) — Cloud Vision 제거
- 지도 = OpenStreetMap — Google Maps 제거
- DeepL Pro (또는 woorichat 자산 활용)
- 상표 3건 (트리거소프트 / 패스웨이브 / 우리쳇)
- Namecheap (글로벌) + Gabia (국내)
- Contabo (운영 + 클론)
- 노란우산공제
- Feature Flag 시스템
- Stage 1~3 자동화 (출시 후)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


OUT = '/Users/m5pro16/Desktop/pathwave/docs/pathwave_external_services_checklist_v2.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "외부서비스_체크리스트_v2"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="2E2E48")
CAT_FILL    = PatternFill("solid", start_color="EEF1F8")
SUM_FILL    = PatternFill("solid", start_color="FFF4D6")
NEW_FILL    = PatternFill("solid", start_color="DEFBDE")     # 이번 세션 신규
CHG_FILL    = PatternFill("solid", start_color="FFE9C7")     # 이번 세션 변경
DEL_FILL    = PatternFill("solid", start_color="FFD5D5")     # 이번 세션 제거(메모)
THIN = Side(style="thin", color="C8CDDA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cell(cell, *, bold=False, color="000000", align="left", fill=None, money=False):
    cell.font = Font(name=FONT_NAME, bold=bold, color=color, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill: cell.fill = fill
    if money: cell.number_format = '₩#,##0;-;"-"'


headers = ["카테고리", "항목", "용도", "가입 시기",
           "결제 주기", "1회 비용", "월 비용", "연 비용",
           "우선순위", "v2 변경", "비고"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    style_cell(cell, bold=True, color="FFFFFF", align="center", fill=HEADER_FILL)


DATA = [
    # ── 행정·사무 (유지)
    ("행정·사무", "창업보육센터 임대료", "사무 공간 임대", "현재~지속", "월", 0, 0, 0, "필수", "", "입주 단가 입력"),
    ("행정·사무", "창업보육센터 관리비", "공용시설", "현재~지속", "월", 0, 0, 0, "필수", "", "별도 입력"),
    ("행정·사무", "주차비 (원광대 월정액)", "업무 차량", "현재~지속", "월", 0, 0, 0, "필수", "", "월정액"),
    ("행정·사무", "세무사 (청운)", "부가세·법인세 신고", "현재~지속", "분기/연", 0, 0, 0, "필수", "", "협의"),
    ("행정·사무", "자동차 (업무용)", "할부/리스/유류", "현재~지속", "월", 0, 0, 0, "필수", "", "합산 입력"),

    # ── 절세·노후 (유지)
    ("절세·노후", "노란우산공제", "1인사업자 노후·압류보호·종합소득세 600만 공제",
     "출시 +3~6개월", "월", 0, 500000, 6000000, "권장", "",
     "익산 희망장려금 월 1만+전북 월 2만 가능"),

    # ── AI·생산성 (유지)
    ("AI·생산성", "Claude (Pro)", "AI 개발·문서", "현재~", "월", 0, 30000, 360000, "필수", "", "$20/월"),
    ("AI·생산성", "ChatGPT (Plus)", "AI 마케팅", "현재~", "월", 0, 30000, 360000, "선택", "", "$20/월"),
    ("AI·생산성", "Gemini (Advanced)", "AI 보완", "출시 후", "월", 0, 30000, 360000, "선택", "", "$20/월"),
    ("AI·생산성", "ADOBE Creative Cloud", "디자인·ASO", "T-7", "월", 0, 72600, 871200, "필수", "", "All Apps"),

    # ── 도메인 (Namecheap 단일 + Gabia)
    ("도메인", "Namecheap (.com/.io)", "글로벌 브랜드", "T-8", "연", 15000, 0, 15000, "필수", "변경", "TLD 미정"),
    ("도메인", "Gabia (.co.kr)", "국내 브랜드", "T-8", "연", 15000, 0, 15000, "필수", "변경", "Namecheap 와 묶음"),

    # ── 서버·호스팅 (Contabo 신규)
    ("서버·호스팅", "Contabo VPS (운영)", "백엔드 운영", "T-3", "월", 20000, 12000, 164000, "필수", "변경", "싱가포르. 한국 리전 검토 가능"),
    ("서버·호스팅", "Contabo VPS (클론)", "Staging — 배포 전 검증", "T-3", "월", 20000, 12000, 164000, "필수", "신규",
     "운영·클론 분리 정책 (사용자 결정 2026-06-05)"),
    ("서버·호스팅", "Vercel (Hobby)", "admin/provider 호스팅", "T-3", "월", 0, 0, 0, "선택", "", "무료"),
    ("서버·호스팅", "Cloudflare Tunnel", "외부 노출", "T-3", "—", 0, 0, 0, "선택", "", "무료"),

    # ── 개발자 계정
    ("개발자 계정", "DUNS Number", "Apple 법인 가입", "T-8", "1회", 0, 0, 0, "필수", "", "무료"),
    ("개발자 계정", "Apple Developer", "iOS App Store", "T-7", "연", 140000, 0, 140000, "필수", "", "$99/년"),
    ("개발자 계정", "Google Play Console", "Android 스토어", "T-7", "1회", 35000, 0, 35000, "필수", "", "$25 일회"),

    # ── Google (Cloud Vision / Maps 제거)
    ("Google", "Workspace Business Starter", "support@ 이메일", "T-8", "월", 0, 10000, 120000, "필수", "", "$7.20/월. Apple Dev 등록 필요"),
    ("Google", "Firebase (Spark)", "FCM 푸시 + 소셜 인증", "T-5", "월", 0, 0, 0, "필수", "", "무료"),
    ("Google", "❌ Google Maps Platform", "(제거됨 — OpenStreetMap 대체)", "—", "—", 0, 0, 0, "X", "제거",
     "이번 세션 결정: Google Maps 비용 회피 → OSM 사용"),
    ("Google", "❌ Google Cloud Vision", "(제거됨 — 디바이스 OCR 대체)", "—", "—", 0, 0, 0, "X", "제거",
     "이번 세션 결정: Apple Vision + ML Kit 사용"),

    # ── 지도·OCR (신규 — 모두 비용 0)
    ("지도·OCR", "OpenStreetMap + flutter_map", "mobile 매장 지도", "T-3", "—", 0, 0, 0, "필수", "신규", "비용 0 — 사용자 정책"),
    ("지도·OCR", "OSM Nominatim", "provider 주소 → 좌표", "T-3", "—", 0, 0, 0, "필수", "신규", "Rate Limit 1 req/sec — 캐시"),
    ("지도·OCR", "Apple Vision Framework", "iOS 메뉴 OCR", "T-2", "—", 0, 0, 0, "필수", "신규", "iOS 단말 처리. Apple Dev 포함"),
    ("지도·OCR", "Google ML Kit Text Recognition", "Android 메뉴 OCR", "T-2", "—", 0, 0, 0, "필수", "신규", "단말 처리. 무료"),

    # ── 결제 PG (제로페이 + 토스 폴백)
    ("결제 PG", "토스페이먼츠", "SP 구독료 + 매장 결제(P2) 폴백", "T-6 (심사 1~2주)", "거래수수료", 0, 0, 0, "필수", "변경",
     "1차 = SP 구독. ~2.9% + ₩30/건"),
    ("결제 PG", "제로페이", "매장 결제 1차 (수수료 0%)", "개발 완료 후 연계", "거래수수료", 0, 0, 0, "권장", "신규",
     "ZeropayProvider 코드 골격 완료. 가맹 후 키 주입"),

    # ── 이메일·알림
    ("이메일·알림", "SendGrid (Essentials)", "회원 이메일 발송", "T-5", "월", 0, 27000, 324000, "필수", "", "$19.95/월"),
    ("이메일·알림", "APNs (.p8)", "iOS 푸시", "T-5", "—", 0, 0, 0, "필수", "", "Apple Dev 포함"),

    # ── 번역 (woorichat 자산 활용 시 0)
    ("번역", "DeepL Pro Advanced", "23 언어 채팅·메뉴 번역", "T-4", "월", 0, 35000, 420000, "필수", "변경",
     "200만자/월. ⚠️ woorichat AI 번역 서버 활용 시 0 (2주 후 사양 확인)"),
    ("번역", "woorichat AI 자동번역 서버 (선택)", "DeepL Pro 대체 검토",
     "검증 후 결정", "정액", 0, 0, 0, "검토", "신규",
     "Contabo 싱가포르. 월 30만원 정액 (사용자 보유)"),

    # ── 모니터링
    ("모니터링", "Sentry (Team)", "크래시·에러 추적", "T-4", "월", 0, 36000, 432000, "필수", "", "$26/월"),

    # ── 상표 등록 (신규 — 사용자 결정)
    ("상표", "상표 #1 트리거소프트 (회사)", "특허청 출원·등록", "출시 전", "1회", 800000, 0, 800000, "필수", "신규",
     "출원료 ₩62k + 등록료 ₩211k + 변리사비 ~₩50만"),
    ("상표", "상표 #2 PathWave (서비스)", "특허청 출원·등록", "출시 전", "1회", 800000, 0, 800000, "필수", "신규", "동일 절차"),
    ("상표", "상표 #3 woorichat (서비스)", "특허청 출원·등록", "출시 전", "1회", 800000, 0, 800000, "필수", "신규",
     "별도 서비스라 별도 출원"),

    # ── 행정 신고 (오늘 가이드 작성)
    ("행정 신고", "통신판매업 신고", "정부24", "사업자등록 + PG 가맹 후", "1회", 40000, 0, 40000, "필수", "신규",
     "구매안전서비스 이용확인증 필수"),
    ("행정 신고", "위치기반서비스 사업 신고", "방통위(KCC) 전자민원", "출시 전", "1회", 0, 0, 0, "필수", "신규",
     "무료. 첨부 서류 6종 (보호조치 명세 등)"),

    # ── 자동화 (출시 후)
    ("자동화 Stage 1", "카톡 챗봇 + 이메일 AI", "매출 자동화", "출시 +1~3개월", "월", 0, 100000, 1200000, "선택", "", "검증 후"),
    ("자동화 Stage 2", "AI 음성통화 + SNS 자동", "마케팅 자동화", "출시 +3~6개월", "월", 0, 250000, 3000000, "선택", "", "Stage 1 안정 후"),
    ("자동화 Stage 3", "CRM + 광고 자동", "풀스택 자동", "출시 +6~12개월", "월", 0, 500000, 6000000, "선택", "", "매출 안정 후"),
]


row_idx = 2
last_cat = None
for d in DATA:
    cat = d[0]
    is_new_cat = (cat != last_cat)
    last_cat = cat
    for col_idx, val in enumerate(d, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        is_money = col_idx in (6, 7, 8)
        align = "right" if is_money else ("center" if col_idx in (1, 4, 5, 9, 10) else "left")

        # 우선순위 색
        priority_color = "000000"
        bold = False
        if col_idx == 9:
            if val == "필수":
                priority_color = "C00000"; bold = True
            elif val == "권장":
                priority_color = "B45F06"; bold = True
            elif val == "선택":
                priority_color = "707070"
            elif val == "X":
                priority_color = "707070"
            elif val == "검토":
                priority_color = "2E74B5"; bold = True

        # v2 변경 컬럼 색
        change_fill = None
        if col_idx == 10:
            if val == "신규":
                change_fill = NEW_FILL; bold = True; priority_color = "0F7B0F"
            elif val == "변경":
                change_fill = CHG_FILL; bold = True; priority_color = "B45F06"
            elif val == "제거":
                change_fill = DEL_FILL; bold = True; priority_color = "C00000"

        # 카테고리 셀
        fill = CAT_FILL if (col_idx == 1 and is_new_cat) else change_fill
        if col_idx == 1 and not is_new_cat:
            cell.value = ""

        style_cell(cell, bold=bold or (col_idx == 1 and is_new_cat),
                   color=priority_color, align=align, fill=fill, money=is_money)
    row_idx += 1


# 합계
total_row = row_idx
ws.cell(row=total_row, column=1, value="합계")
ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
for col_idx in (1, 6, 7, 8):
    cell = ws.cell(row=total_row, column=col_idx)
    style_cell(cell, bold=True, align="right" if col_idx > 1 else "center",
               fill=SUM_FILL, money=(col_idx in (6, 7, 8)))
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=total_row, column=6).value = f"=SUM(F2:F{total_row-1})"
ws.cell(row=total_row, column=7).value = f"=SUM(G2:G{total_row-1})"
ws.cell(row=total_row, column=8).value = f"=SUM(H2:H{total_row-1})"
for col_idx in (9, 10, 11):
    cell = ws.cell(row=total_row, column=col_idx)
    style_cell(cell, fill=SUM_FILL)


widths = [16, 32, 32, 22, 12, 12, 12, 12, 10, 10, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 34
for r in range(2, total_row + 1):
    ws.row_dimensions[r].height = 30

ws.freeze_panes = "A2"


# 범례 시트
ws2 = wb.create_sheet("v2 변경 범례")
legend = [
    ["색", "의미", "설명"],
    ["연두", "신규", "이번 세션(2026-06-05~09) 결정으로 새로 추가된 항목"],
    ["주황", "변경", "기존에 있었지만 정책·비용·구조 변경된 항목"],
    ["빨강", "제거", "이번 세션 결정으로 사용 안 하기로 한 항목 (메모용 잔존)"],
    ["—", "유지", "v1 와 동일"],
]
for r, row in enumerate(legend, 1):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        is_header = (r == 1)
        fill = None
        if not is_header and c == 1:
            fill = {1: None, 2: NEW_FILL, 3: CHG_FILL, 4: DEL_FILL, 5: None}.get(r)
        style_cell(cell, bold=is_header,
                   color="FFFFFF" if is_header else "000000",
                   align="center" if c <= 2 else "left",
                   fill=HEADER_FILL if is_header else fill)
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 60


Path(OUT).parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f'생성: {OUT}')
print(f'  카테고리 13개 · 총 {row_idx - 2}개 항목 + 합계')
